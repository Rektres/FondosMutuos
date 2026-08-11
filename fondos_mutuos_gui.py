import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import pandas as pd
from datetime import datetime, timedelta
import threading
import time
import os
import glob
import json
import re
import subprocess
import sys
import tempfile
import urllib.request
import webbrowser
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from tkcalendar import DateEntry

CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".fondos_mutuos_prefs.json")

# Identidad de la aplicación (se muestra en el botón "Información")
APP_NOMBRE = "Descargador de Fondos Mutuos CMF"
APP_VERSION = "1.0"
APP_AUTOR = "Mateo Araneda Medina"
APP_LINKEDIN = "https://www.linkedin.com/in/mateo-a-6388a7219/"          # URL del perfil; si queda vacía no se muestra el enlace
URL_CMF = "https://www.cmfchile.cl/institucional/estadisticas/fondos_cartola_diaria.php"
NOMBRE_LOGO = "Logo.png"

# Indicadores del día (UF, UTM, dólar). API pública, sin credenciales.
URL_INDICADORES = "https://mindicador.cl/api"
INDICADORES = [("uf", "UF", 2), ("utm", "UTM", 0), ("dolar", "Dólar", 2)]
REVISION_INDICADORES_MS = 30 * 60 * 1000     # cada 30 min se revisa si cambió el día


def _dia_habil_vencido(fecha_base):
    """Último día hábil anterior a fecha_base (salta sábado/domingo)."""
    fecha = fecha_base - timedelta(days=1)
    while fecha.weekday() >= 5:
        fecha -= timedelta(days=1)
    return fecha


# Tipografía ampliada para lectura cómoda
FONT_NORMAL = ("Segoe UI", 14)
FONT_BOLD = ("Segoe UI", 14, "bold")
FONT_BOTON = ("Segoe UI", 15, "bold")
ALTO_FILA = 33
ALTO_FILTRO = 32
PASO_MARQUEE = 320       # ms entre cada avance del nombre que se desplaza
VISIBLES_PESTANA = 28    # caracteres del nombre que caben en una pestaña lateral

# Neumorphism Colors
BG_PRIMARY = "#E8EDF3"
BG_SECONDARY = "#F5F7FC"
TEXT_PRIMARY = "#37474F"
TEXT_SECONDARY = "#6B7A89"
ACCENT_LIGHT = "#FFFFFF"
ACCENT_DARK = "#D1D9E6"
CAMPO_COLOR = "#F7F9FC"      # fondo de campos y tabla

# Botones: tonos propios de la paleta, siempre distintos del fondo
BOTON_BG = "#C3D2E3"
BOTON_BG_HOVER = "#AFC3DA"
BOTON_BG_PRESS = "#9BB2CE"
BOTON_TEXTO = "#1F3247"
ACCION_BG = "#3D6C9C"        # acciones principales (contraste >= 4.5:1 con texto blanco)
ACCION_BG_HOVER = "#325A83"
ACCION_TEXTO = "#FFFFFF"

# Columnas de la tabla: (id, etiqueta, campo CSV, ancho, alineación)
COLUMNAS = [
    ("fecha", "Fecha", "FECHA_INF", 135, "w"),
    ("entidad", "Entidad", "NOM_ADM", 400, "w"),
    ("num_fondo", "N° Fondo", "RUN_FM", 125, "w"),
    ("nombre_fondo", "Nombre Fondo", "NOMBRE_FM", 300, "w"),
    ("serie", "Serie", "SERIE", 130, "w"),
    ("cuota", "Valor Cuota", "VALOR_CUOTA", 165, "e"),
]
COL_IDS = [c[0] for c in COLUMNAS]
COL_CAMPO = {c[0]: c[2] for c in COLUMNAS}
COL_ETIQUETA = {c[0]: c[1] for c in COLUMNAS}

# Texto guía dentro de cada caja de filtro de la tabla
PLACEHOLDER_COL = {
    "fecha": "aaaa-mm-dd",
    "entidad": "buscar entidad...",
    "num_fondo": "n° de fondo...",
    "nombre_fondo": "buscar nombre...",
    "serie": "buscar serie...",
    "cuota": "buscar valor...",
}
PLACEHOLDER_BUSCAR = "escribe para buscar..."

# Campos que deben quedar numéricos para poder ordenar y exportar a Excel como número
COLUMNAS_NUMERICAS = ["VALOR_CUOTA"]

# Contenido que se exporta a Excel bajo el encabezado de cada entidad
CAMPOS_EXPORT = [
    ("FECHA_INF", "Fecha"),
    ("RUN_FM", "N° Fondo"),
    ("NOMBRE_FM", "Nombre Fondo"),
    ("SERIE", "Serie"),
    ("VALOR_CUOTA", "Valor Cuota"),
]

# Columnas de la plantilla de carga/descarga
COLS_PLANTILLA = ["Entidad", "N° Fondo", "Nombre Fondo", "Serie"]

# Valor del dropdown de serie que significa "no acotar"
TODAS = "Todas"


def _clave_fondo(run):
    """Ordena los fondos numéricamente cuando el RUN es numérico."""
    texto = str(run)
    return (0, int(texto), "") if texto.isdigit() else (1, 0, texto)


def _acortar(texto, limite):
    """Recorta para que la etiqueta no empuje al dropdown fuera de la vista."""
    texto = str(texto)
    return texto if len(texto) <= limite else texto[:limite - 1] + "…"


def _desborda(canvas):
    """True si el contenido supera el alto visible, o sea si hay algo que desplazar."""
    caja = canvas.bbox("all")
    if not caja:
        return False
    return (caja[3] - caja[1]) > canvas.winfo_height()


def _ruta_recurso(nombre):
    """Ruta de un archivo que acompaña al programa, ya sea .py suelto o .exe."""
    base = getattr(sys, "_MEIPASS", os.path.dirname(os.path.abspath(__file__)))
    ruta = os.path.join(base, nombre)
    if os.path.exists(ruta):
        return ruta
    # Al ejecutar el .py desde otra carpeta, se busca junto al propio archivo
    alterna = os.path.join(os.path.dirname(os.path.abspath(__file__)), nombre)
    return alterna if os.path.exists(alterna) else None


def _cargar_logo(lado, atenuar=None, fondo=None):
    """Devuelve el logo recortado en CÍRCULO, como PhotoImage de `lado` px.

    El PNG original trae el motivo circular sobre un fondo negro rectangular; se
    recorta el disco y lo de afuera queda transparente, sin esquinas oscuras.
    Con `atenuar` (0-1) se mezcla contra `fondo` para usarlo de marca de agua.
    Devuelve None si no se puede cargar.
    """
    ruta = _ruta_recurso(NOMBRE_LOGO)
    if not ruta:
        return None
    try:
        from PIL import Image, ImageDraw, ImageTk

        imagen = Image.open(ruta).convert("RGB")
        # El motivo está centrado: se toma el cuadrado del medio
        ancho, alto = imagen.size
        corte = min(ancho, alto)
        izq, arriba = (ancho - corte) // 2, (alto - corte) // 2
        imagen = imagen.crop((izq, arriba, izq + corte, arriba + corte))

        # Se trabaja en grande y se reduce al final, para que el borde salga suave
        trabajo = max(lado * 4, 256)
        imagen = imagen.resize((trabajo, trabajo), Image.LANCZOS)

        if atenuar is not None and fondo:
            plano = Image.new("RGB", imagen.size, fondo)
            imagen = Image.blend(plano, imagen, atenuar)

        # Máscara circular: fuera del disco, transparente
        mascara = Image.new("L", (trabajo, trabajo), 0)
        ImageDraw.Draw(mascara).ellipse((0, 0, trabajo - 1, trabajo - 1), fill=255)

        imagen = imagen.convert("RGBA")
        imagen.putalpha(mascara)
        imagen = imagen.resize((lado, lado), Image.LANCZOS)
        return ImageTk.PhotoImage(imagen)
    except Exception:
        pass
    try:
        # Respaldo sin Pillow: solo reducción por factor entero, sin recorte circular
        original = tk.PhotoImage(file=ruta)
        factor = max(1, min(original.width(), original.height()) // max(lado, 1))
        return original.subsample(factor, factor)
    except Exception:
        return None


def _fecha_corta(iso):
    """'2026-08-10' -> '10/08/2026'. Devuelve el original si no calza el formato."""
    texto = str(iso or "")[:10]
    partes = texto.split("-")
    return f"{partes[2]}/{partes[1]}/{partes[0]}" if len(partes) == 3 else texto


def _formato_clp(valor, decimales=2):
    """Formatea a la chilena: punto para los miles y coma para los decimales."""
    texto = f"{float(valor):,.{decimales}f}"
    return texto.replace(",", "\x00").replace(".", ",").replace("\x00", ".")


def _consultar_indicadores():
    """Trae UF, UTM y dólar del día. Devuelve {} si no se puede consultar."""
    try:
        peticion = urllib.request.Request(
            URL_INDICADORES, headers={"User-Agent": f"{APP_NOMBRE}/{APP_VERSION}"})
        with urllib.request.urlopen(peticion, timeout=12) as respuesta:
            datos = json.loads(respuesta.read().decode("utf-8"))
    except Exception:
        return {}                      # sin conexión o servicio caído

    salida = {}
    for clave, _etiqueta, _dec in INDICADORES:
        bloque = datos.get(clave)
        if isinstance(bloque, dict) and bloque.get("valor") is not None:
            salida[clave] = bloque["valor"]
            fecha = str(bloque.get("fecha", ""))[:10]
            if fecha:
                salida.setdefault("fechas", {})[clave] = fecha
    return salida


def _abrir_carpeta_de(archivo):
    """Abre el explorador en la carpeta del archivo, dejándolo seleccionado."""
    ruta = os.path.abspath(archivo)
    try:
        # /select deja el archivo resaltado dentro de su carpeta
        subprocess.Popen(["explorer", "/select,", ruta])
        return True
    except Exception:
        pass
    try:
        os.startfile(os.path.dirname(ruta))    # respaldo: solo abrir la carpeta
        return True
    except Exception:
        return False


def _poner_placeholder(entry, variable, texto):
    """Muestra un texto guía en gris mientras el campo está vacío y sin foco."""
    def mostrar():
        if not variable.get():
            entry._placeholder_activo = True
            variable.set(texto)
            entry.configure(foreground=TEXT_SECONDARY)

    def limpiar(_evento=None):
        if getattr(entry, "_placeholder_activo", False):
            entry._placeholder_activo = False
            variable.set("")
            entry.configure(foreground=TEXT_PRIMARY)

    entry.bind("<FocusIn>", limpiar)
    entry.bind("<FocusOut>", lambda _e: mostrar())
    entry._mostrar_placeholder = mostrar
    mostrar()
    # Se compara el contenido, no un flag de foco: así también funciona si la
    # variable se cambia por código (reset de filtros, pruebas, plantillas)
    return lambda: "" if variable.get() == texto else variable.get()


def _normalizar_numericas(df):
    """La CMF entrega los decimales con punto; fuerza tipo numérico en esas columnas."""
    for col in COLUMNAS_NUMERICAS:
        if col not in df.columns or pd.api.types.is_numeric_dtype(df[col]):
            continue
        serie = df[col].astype(str).str.strip()
        convertida = pd.to_numeric(serie, errors="coerce")
        if convertida.isna().all() and not serie.empty:
            # Respaldo por si algún día el archivo llegara con coma decimal
            convertida = pd.to_numeric(
                serie.str.replace(".", "", regex=False).str.replace(",", ".", regex=False),
                errors="coerce")
        df[col] = convertida
    return df


def _num(valor, defecto=0.0):
    """Convierte a float tolerando vacíos, nulos y texto no numérico."""
    try:
        v = float(valor)
    except (TypeError, ValueError):
        return defecto
    return defecto if pd.isna(v) else v


def _extraer_catalogo_fondos(driver):
    """Lee del formulario de la CMF las opciones 'NUMERO-NOMBRE' de cada fondo.

    Se traen los 1.500+ textos en UNA sola llamada al navegador. Recorrerlos con
    find_elements + .text costaba un viaje de ida y vuelta por opción y hacía que
    la ventana tardara segundos en quedar lista.
    """
    catalogo = {}
    try:
        textos = driver.execute_script(
            "return Array.from(document.querySelectorAll('option'))"
            ".map(function(o){ return o.textContent; });") or []
        for texto in textos:
            # textContent llega crudo: se colapsan los espacios repetidos para que
            # los nombres queden como los muestra el navegador
            limpio = re.sub(r"\s+", " ", texto or "").strip()
            coincide = re.match(r"^(\d+)\s*-\s*(.+)$", limpio)
            if coincide:
                catalogo[coincide.group(1)] = coincide.group(2).strip()
    except Exception:
        pass
    return catalogo


# JS que deja a la vista solo la rama del formulario (captcha + botón), ocultando
# menús, banners y encuestas del portal. No mueve nodos: el script del sitio que
# valida el captcha y hace el submit sigue funcionando igual.
JS_SOLO_FORMULARIO = """
var form = document.form1 || document.querySelector('form[name="form1"]')
           || (document.querySelector('#captcha_img') || {}).form;
if (!form) { return 'sin-formulario'; }

/* 1) Ocultar todo lo que quede fuera de la rama del formulario (menús, banners) */
var conservar = new Set();
for (var n = form; n; n = n.parentElement) { conservar.add(n); }
conservar.forEach(function (nodo) {
    if (!nodo.parentElement) { return; }
    Array.prototype.forEach.call(nodo.parentElement.children, function (hijo) {
        if (!conservar.has(hijo) && hijo.tagName !== 'SCRIPT'
            && hijo.tagName !== 'STYLE' && hijo.tagName !== 'LINK') {
            hijo.style.display = 'none';
        }
    });
});

/* 2) Dentro del formulario, quitar los textos explicativos largos: sin ellos el
      captcha y el botón quedan a la vista sin necesidad de desplazarse */
var bloques = form.querySelectorAll('p,div,table,h1,h2,h3,h4,ul,ol,li,span,center,blockquote');
Array.prototype.forEach.call(bloques, function (e) {
    if (e.querySelector('input,select,img,button,textarea')) { return; }
    if ((e.textContent || '').trim().length > 55) { e.style.display = 'none'; }
});

var estilo = document.createElement('style');
estilo.textContent =
    'html,body{background:#E8EDF3 !important;margin:0 !important;padding:0 !important;}' +
    'body *{max-width:100% !important;}' +
    '#cmf_ayuda{font:600 16px "Segoe UI",sans-serif;color:#1F3247;background:#C3D2E3;' +
    'padding:9px 14px;margin:0 0 6px 0;line-height:1.35;}' +
    'input[type=text]{font-size:17px !important;padding:6px !important;}' +
    'img#captcha_img{transform:scale(1.35);transform-origin:left center;margin:10px 0;}';
document.head.appendChild(estilo);

/* 3) Instrucción propia, justo encima del formulario */
var ayuda = document.createElement('div');
ayuda.id = 'cmf_ayuda';
ayuda.innerHTML = 'Las fechas ya están puestas.<br>' +
                  '1) Escriba los caracteres de la imagen &nbsp;&nbsp; ' +
                  '2) Presione GENERAR ARCHIVO';
form.insertBefore(ayuda, form.firstChild);   // dentro del form: viaja con él al hacer scroll
form.scrollIntoView({block: 'start'});

/* 4) Dejar el cursor puesto en el campo del captcha: el usuario solo escribe.
      Enter equivale a presionar GENERAR ARCHIVO. */
var campoCaptcha = document.querySelector('input[name=captcha]');
if (campoCaptcha) {
    campoCaptcha.focus();
    campoCaptcha.select();
    campoCaptcha.addEventListener('keydown', function (ev) {
        if (ev.key === 'Enter') {
            ev.preventDefault();
            var b = document.querySelector('input[name=btnConsulta], input[name=Submit]');
            if (b) { b.click(); }
        }
    });
}
return 'ok';
"""


def _agregar_nombre_fondo(df, catalogo):
    """Agrega la columna NOMBRE_FM cruzando RUN_FM contra el catálogo de la CMF."""
    if "RUN_FM" in df.columns:
        df["NOMBRE_FM"] = df["RUN_FM"].astype(str).map(lambda r: catalogo.get(r, ""))
    return df


class FondosMutuosApp:
    def __init__(self, root):
        self.root = root
        self.root.title(APP_NOMBRE)
        self.root.config(bg=BG_PRIMARY)

        # El logo como ícono. Con default=True lo heredan todos los diálogos que
        # se abran después, sin tener que repetirlo en cada uno.
        self._icono = _cargar_logo(64)
        if self._icono is not None:
            try:
                self.root.iconphoto(True, self._icono)
            except tk.TclError:
                pass
        self.df = None
        self.download_dir = os.path.join(tempfile.gettempdir(), "cmf_fondos_mutuos_downloads")
        os.makedirs(self.download_dir, exist_ok=True)

        style = ttk.Style()
        style.theme_use("clam")

        # Acciones principales: color propio y texto claro
        style.configure("Big.TButton",
                       padding=(20, 22),
                       font=FONT_BOTON,
                       background=ACCION_BG,
                       foreground=ACCION_TEXTO,
                       borderwidth=0,
                       relief="flat",
                       focuscolor=ACCION_BG)
        style.map("Big.TButton",
                 background=[("active", ACCION_BG_HOVER), ("pressed", ACCION_BG_HOVER)],
                 foreground=[("active", ACCION_TEXTO), ("pressed", ACCION_TEXTO)])

        # Botones normales: tono de la paleta, siempre distinto del fondo
        style.configure("TButton",
                       padding=(12, 9),
                       font=FONT_NORMAL,
                       background=BOTON_BG,
                       foreground=BOTON_TEXTO,
                       borderwidth=0,
                       relief="flat",
                       focuscolor=BOTON_BG)
        style.map("TButton",
                 background=[("active", BOTON_BG_HOVER), ("pressed", BOTON_BG_PRESS)],
                 foreground=[("active", BOTON_TEXTO), ("pressed", BOTON_TEXTO)])

        # Neumorphic label style
        style.configure("TLabel",
                       font=FONT_NORMAL,
                       background=BG_PRIMARY,
                       foreground=TEXT_PRIMARY)

        # Neumorphic frame style
        style.configure("TFrame",
                       background=BG_PRIMARY)

        # Neumorphic combobox style
        style.configure("TCombobox",
                       font=FONT_NORMAL,
                       background=CAMPO_COLOR,
                       foreground=TEXT_PRIMARY,
                       fieldbackground=CAMPO_COLOR)
        self.root.option_add("*TCombobox*Listbox.font", FONT_NORMAL)
        self.root.option_add("*TCombobox*Listbox.background", CAMPO_COLOR)
        self.root.option_add("*TCombobox*Listbox.foreground", TEXT_PRIMARY)

        # Neumorphic entry style
        style.configure("TEntry",
                       background=CAMPO_COLOR,
                       foreground=TEXT_PRIMARY,
                       fieldbackground=CAMPO_COLOR,
                       borderwidth=1)

        # Neumorphic checkbutton style
        style.configure("TCheckbutton",
                       background=BG_PRIMARY,
                       foreground=TEXT_PRIMARY,
                       font=FONT_NORMAL,
                       indicatorsize=18,
                       padding=(4, 4))

        # Barras de desplazamiento con contraste suficiente para verse
        style.configure("TScrollbar",
                       background=BOTON_BG,
                       troughcolor=BG_SECONDARY,
                       bordercolor=BG_SECONDARY,
                       arrowcolor=BOTON_TEXTO,
                       width=16)
        style.map("TScrollbar", background=[("active", BOTON_BG_HOVER)])

        # Pestañas laterales por entidad (una fila por entidad, nombre completo legible)
        style.configure("Tab.TButton",
                       anchor="w",
                       padding=(10, 9),
                       font=FONT_NORMAL,
                       background=BOTON_BG,
                       foreground=BOTON_TEXTO,
                       borderwidth=0,
                       relief="flat")
        style.map("Tab.TButton",
                 background=[("active", BOTON_BG_HOVER), ("pressed", BOTON_BG_PRESS)],
                 foreground=[("active", BOTON_TEXTO), ("pressed", BOTON_TEXTO)])
        style.configure("TabSel.TButton",
                       anchor="w",
                       padding=(10, 9),
                       font=FONT_BOLD,
                       background=ACCION_BG,
                       foreground=ACCION_TEXTO,
                       borderwidth=0,
                       relief="flat")
        style.map("TabSel.TButton",
                 background=[("active", ACCION_BG), ("pressed", ACCION_BG)],
                 foreground=[("active", ACCION_TEXTO), ("pressed", ACCION_TEXTO)])

        hasta_default = _dia_habil_vencido(datetime.now())
        desde_default = hasta_default

        # Panel superior: fechas y acciones
        frame_top = ttk.Frame(root, padding="15")
        frame_top.pack(fill=tk.X)

        ttk.Label(frame_top, text="Desde:").pack(side=tk.LEFT, padx=5)
        self.entry_desde = DateEntry(frame_top, width=11, date_pattern="yyyy-mm-dd", font=FONT_NORMAL,
                                      background=CAMPO_COLOR, foreground=TEXT_PRIMARY, borderwidth=1,
                                      maxdate=hasta_default)
        self.entry_desde.pack(side=tk.LEFT, padx=5)
        self.entry_desde.set_date(desde_default)

        ttk.Label(frame_top, text="Hasta:").pack(side=tk.LEFT, padx=5)
        self.entry_hasta = DateEntry(frame_top, width=11, date_pattern="yyyy-mm-dd", font=FONT_NORMAL,
                                      background=CAMPO_COLOR, foreground=TEXT_PRIMARY, borderwidth=1,
                                      maxdate=hasta_default)
        self.entry_hasta.pack(side=tk.LEFT, padx=5)
        self.entry_hasta.set_date(hasta_default)

        ttk.Button(frame_top, text="Descargar datos", style="Big.TButton",
                   command=self.descargar_datos).pack(side=tk.LEFT, padx=10, fill=tk.BOTH, expand=True)
        ttk.Button(frame_top, text="Exportar Excel", style="Big.TButton",
                   command=self.exportar_excel).pack(side=tk.LEFT, padx=10, fill=tk.BOTH, expand=True)

        self.label_status = ttk.Label(frame_top, text="", foreground=TEXT_SECONDARY)
        self.label_status.pack(side=tk.LEFT, padx=20)

        # Panel de configuración
        frame_config = ttk.Frame(root, padding="10")
        frame_config.pack(fill=tk.X, padx=5)
        ttk.Button(frame_config, text="⚙ Preferencias",
                   command=self._abrir_preferencias).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_config, text="⬇ Descargar plantilla",
                   command=self.descargar_plantilla).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_config, text="⬆ Cargar plantilla",
                   command=self.cargar_plantilla).pack(side=tk.LEFT, padx=5)
        ttk.Button(frame_config, text="ℹ Información",
                   command=self._abrir_informacion).pack(side=tk.LEFT, padx=5)

        # Indicadores del día, al costado del botón de información: los valores
        # arriba y la fecha a la que corresponde cada uno debajo
        self.frame_indicadores = tk.Frame(frame_config, bg=BOTON_BG, padx=12, pady=6,
                                          cursor="hand2")
        self.frame_indicadores.pack(side=tk.LEFT, padx=(18, 5))
        self.label_indicadores = tk.Label(self.frame_indicadores,
                                          text="UF · UTM · Dólar: cargando…",
                                          bg=BOTON_BG, fg=BOTON_TEXTO, font=FONT_BOLD,
                                          cursor="hand2")
        self.label_indicadores.pack(anchor=tk.W)
        self.label_ind_fechas = tk.Label(self.frame_indicadores, text="", bg=BOTON_BG,
                                         fg=BOTON_TEXTO, font=("Segoe UI", 11),
                                         cursor="hand2")
        self.label_ind_fechas.pack(anchor=tk.W)
        for w in (self.frame_indicadores, self.label_indicadores, self.label_ind_fechas):
            w.bind("<Button-1>", lambda _e: self._refrescar_indicadores(True))

        self.label_plantilla = ttk.Label(frame_config, text="", foreground=TEXT_SECONDARY)
        self.label_plantilla.pack(side=tk.LEFT, padx=20)

        # Panel de filtros: entidad, búsqueda y fechas
        frame_filtros = ttk.Frame(root, padding="15")
        frame_filtros.pack(fill=tk.X)

        self.entidades_disponibles = []
        self.entidad_vars = {}
        self.popup_entidades = None
        self.btn_entidad = ttk.Button(frame_filtros, text="Entidad ▾", command=self._abrir_selector_entidades)
        self.btn_entidad.pack(side=tk.LEFT, padx=5)

        self.fondos_disponibles = []
        self.fondo_vars = {}
        self.fondo_entidad = {}
        self.fondo_nombre = {}
        self.series_por_fondo = {}   # run -> lista de series disponibles
        self.fondo_series = {}       # run -> conjunto de series marcadas
        self.popup_fondos = None
        self._refrescar_tabs_fondos = None
        self._marquee = []              # nombres largos que se desplazan
        self._marquee_activo = False
        self._marquee_after = None      # id del único temporizador de la animación
        self.btn_fondo = ttk.Button(frame_filtros, text="Fondos ▾", command=self._abrir_selector_fondos)
        self.btn_fondo.pack(side=tk.LEFT, padx=5)

        # DateEntry no emite <<Change>>; el trace sobre su textvariable sí capta
        # tanto la elección en el calendario como la escritura a mano
        ttk.Label(frame_filtros, text="Fecha desde:").pack(side=tk.LEFT, padx=(20, 5))
        hoy = datetime.now().date()
        self.var_fecha_desde = tk.StringVar()
        self._date_desde_filter = DateEntry(frame_filtros, width=11, date_pattern="yyyy-mm-dd",
                                            font=FONT_NORMAL, textvariable=self.var_fecha_desde,
                                            background=CAMPO_COLOR, foreground=TEXT_PRIMARY,
                                            borderwidth=1, maxdate=hoy)
        self._date_desde_filter.pack(side=tk.LEFT, padx=5)
        self._date_desde_filter.set_date(hoy)
        self.var_fecha_desde.trace_add("write", self._aplicar_filtros)

        ttk.Label(frame_filtros, text="Hasta:").pack(side=tk.LEFT, padx=5)
        self.var_fecha_hasta = tk.StringVar()
        self._date_hasta_filter = DateEntry(frame_filtros, width=11, date_pattern="yyyy-mm-dd",
                                            font=FONT_NORMAL, textvariable=self.var_fecha_hasta,
                                            background=CAMPO_COLOR, foreground=TEXT_PRIMARY,
                                            borderwidth=1, maxdate=hoy)
        self._date_hasta_filter.pack(side=tk.LEFT, padx=5)
        self._date_hasta_filter.set_date(hoy)
        self.var_fecha_hasta.trace_add("write", self._aplicar_filtros)

        self.label_seleccion = ttk.Label(frame_filtros, text="", foreground=TEXT_SECONDARY)
        self.label_seleccion.pack(side=tk.LEFT, padx=(20, 5))

        # Tabla
        frame_tabla = ttk.Frame(root, padding="10")
        frame_tabla.pack(fill=tk.BOTH, expand=True)

        style.configure("Treeview",
                       font=FONT_NORMAL,
                       rowheight=ALTO_FILA,
                       background=CAMPO_COLOR,
                       foreground=TEXT_PRIMARY,
                       fieldbackground=CAMPO_COLOR,
                       borderwidth=1)
        style.configure("Treeview.Heading",
                       font=FONT_BOLD,
                       background=BG_SECONDARY,
                       foreground=TEXT_PRIMARY,
                       borderwidth=1)
        style.map("Treeview",
                 background=[("selected", ACCENT_DARK)],
                 foreground=[("selected", TEXT_PRIMARY)])

        # Fila de filtros, uno por columna, alineada con las cabeceras
        self.frame_filtros_col = ttk.Frame(frame_tabla, height=ALTO_FILTRO + 2)
        self.frame_filtros_col.pack(fill=tk.X)
        self.frame_filtros_col.pack_propagate(False)

        frame_tree = ttk.Frame(frame_tabla)
        frame_tree.pack(fill=tk.BOTH, expand=True)

        scrollbar = ttk.Scrollbar(frame_tree)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.tree = ttk.Treeview(frame_tree, yscrollcommand=scrollbar.set, height=20,
                                 selectmode="extended", show="headings")
        scrollbar.config(command=self.tree.yview)

        self.tree["columns"] = tuple(COL_IDS)
        for col_id, etiqueta, _campo, ancho, anchor in COLUMNAS:
            # Entidad se dimensiona al contenido; el sobrante lo absorbe Nombre Fondo
            self.tree.column(col_id, anchor=anchor, width=ancho, minwidth=ancho,
                             stretch=(col_id == "nombre_fondo"))
            self.tree.heading(col_id, text=etiqueta, anchor=anchor,
                              command=lambda c=col_id: self._ordenar_por_columna(c))

        self.tree.pack(fill=tk.BOTH, expand=True)

        # Entradas de filtro por columna
        self.filtros_columna = {}
        self.filtro_entries = {}
        self.filtro_lectores = {}
        for col_id in COL_IDS:
            var = tk.StringVar()
            entry = ttk.Entry(self.frame_filtros_col, textvariable=var, font=FONT_NORMAL)
            self.filtros_columna[col_id] = var
            self.filtro_entries[col_id] = entry
            self.filtro_lectores[col_id] = _poner_placeholder(
                entry, var, PLACEHOLDER_COL.get(col_id, "buscar..."))
            var.trace_add("write", self._aplicar_filtros)
        self.tree.bind("<Configure>", lambda e: self.root.after_idle(self._reposicionar_filtros_columna))
        self.root.after(100, self._reposicionar_filtros_columna)

        # Estado de ordenamiento
        self._sort_col = None
        self._sort_asc = True

        # Selección por celda (rango rectangular filas x columnas)
        self._sel_ancla = None
        self._sel_filas = []
        self._sel_cols = []
        self.tree.bind("<Button-1>", self._on_tree_press)
        self.tree.bind("<B1-Motion>", self._on_tree_drag)
        self.tree.bind("<ButtonRelease-1>", self._on_tree_release)

        # Copiar celdas seleccionadas
        self.tree.bind("<Control-c>", self._copiar_seleccion)
        self.menu_contextual = tk.Menu(self.root, tearoff=0,
                                       bg=CAMPO_COLOR, fg=TEXT_PRIMARY,
                                       activebackground=ACCENT_DARK,
                                       activeforeground=TEXT_PRIMARY,
                                       borderwidth=1)
        self.menu_contextual.add_command(label="Copiar selección", command=self._copiar_seleccion)
        self.tree.bind("<Button-3>", self._mostrar_menu_contextual)

        self._centrar_ventana(1560, 900)
        # Se arranca aquí y solo aquí: un temporizador para toda la vida de la app
        self._marquee_after = self.root.after(PASO_MARQUEE, self._animar_marquee)

        # Indicadores del día: se piden al abrir y se revisan por si cambia la fecha
        self._indicadores_dia = None
        self._pendiente_indicadores = None
        self.root.after(200, self._refrescar_indicadores)
        self.root.after(REVISION_INDICADORES_MS, self._revisar_cambio_de_dia)

    def _centrar_ventana(self, ancho, alto):
        self.root.update_idletasks()
        ancho = min(ancho, self.root.winfo_screenwidth() - 40)
        alto = min(alto, self.root.winfo_screenheight() - 80)
        x = (self.root.winfo_screenwidth() - ancho) // 2
        y = (self.root.winfo_screenheight() - alto) // 2
        self.root.geometry(f"{ancho}x{alto}+{x}+{y}")

    def _area_scroll(self, contenedor, alto=None):
        """Crea un canvas con scroll vertical y devuelve (canvas, frame interior)."""
        canvas = tk.Canvas(contenedor, highlightthickness=0, bg=BG_PRIMARY,
                           highlightbackground=BG_PRIMARY,
                           **({"height": alto} if alto else {}))
        barra = ttk.Scrollbar(contenedor, orient="vertical", command=canvas.yview)
        interior = ttk.Frame(canvas)
        ventana = canvas.create_window((0, 0), window=interior, anchor="nw")

        def _ajustar(_evento=None):
            """Recalcula el área desplazable; si todo cabe, devuelve la vista al tope.

            Sin el moveto(0) la lista se podía 'empujar' hacia abajo aunque quedara
            una sola opción visible, dejándola fuera de la vista.
            """
            canvas.configure(scrollregion=canvas.bbox("all"))
            if not _desborda(canvas):
                canvas.yview_moveto(0)

        interior.bind("<Configure>", _ajustar)
        # El interior sigue el ancho del canvas para que nada quede fuera de la vista
        canvas.bind("<Configure>", lambda e: (canvas.itemconfigure(ventana, width=e.width),
                                              _ajustar()))
        canvas.configure(yscrollcommand=barra.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        barra.pack(side=tk.RIGHT, fill=tk.Y)
        canvas._ajustar_scroll = _ajustar        # para refrescar después de filtrar
        return canvas, interior

    def _registrar_marquee(self, widget, prefijo, nombre, visibles, animar=False):
        """Deja el texto fijo, y solo si `animar` y el nombre no cabe, lo pone en marcha.

        Se anima únicamente la pestaña seleccionada: con 20+ entidades moviéndose a la
        vez el diálogo resulta ilegible. Aquí solo se toca la lista — el temporizador
        es único y vive aparte, para que reregistrar no multiplique la velocidad.
        """
        self._marquee = [r for r in self._marquee if r["widget"] is not widget]
        widget._marquee_nombre = nombre          # el nombre íntegro, para consultarlo
        if len(nombre) <= visibles or not animar:
            widget.configure(text=prefijo + _acortar(nombre, visibles))
            self._marquee_activo = bool(self._marquee)
            return
        self._marquee.append({"widget": widget, "prefijo": prefijo, "pos": 0,
                              "ciclo": nombre + "     ·     ", "visibles": visibles})
        widget.configure(text=prefijo + nombre[:visibles])
        self._marquee_activo = True

    def _paso_marquee(self):
        """Avanza un paso los nombres en movimiento. No programa nada."""
        vivos = []
        for reg in self._marquee:
            if not reg["widget"].winfo_exists():
                continue
            vivos.append(reg)
            reg["pos"] = (reg["pos"] + 1) % len(reg["ciclo"])
            doble = reg["ciclo"] + reg["ciclo"]
            reg["widget"].configure(
                text=reg["prefijo"] + doble[reg["pos"]:reg["pos"] + reg["visibles"]])
        self._marquee = vivos
        self._marquee_activo = bool(vivos)

    def _animar_marquee(self):
        """Tick del bucle: avanza un paso y deja programado exactamente un tick más.

        Cancela siempre el pendiente antes de reprogramar. Sin esto, cualquier camino
        que vuelva a entrar aquí (cambiar de pestaña, una llamada directa) sumaba otro
        temporizador y el nombre se movía cada vez más rápido.
        """
        self._paso_marquee()
        if self._marquee_after is not None:
            try:
                self.root.after_cancel(self._marquee_after)
            except Exception:
                pass
            self._marquee_after = None
        try:
            self._marquee_after = self.root.after(PASO_MARQUEE, self._animar_marquee)
        except tk.TclError:
            self._marquee_after = None      # la ventana ya se cerró

    def _pestanas_por_entidad(self, contenedor, entidades, constructor, ancho_tabs=380,
                              contador=None):
        """Pestañas laterales: una por entidad (nombre completo), contenido a la derecha.

        Se usan pestañas propias en vez de ttk.Notebook porque con muchas entidades
        el Notebook comprime los títulos hasta dejarlos ilegibles.
        """
        marco = ttk.Frame(contenedor)
        marco.pack(fill=tk.BOTH, expand=True)

        col_izq = ttk.Frame(marco, width=ancho_tabs)
        col_izq.pack(side=tk.LEFT, fill=tk.Y)
        col_izq.pack_propagate(False)
        canvas_tabs, lista_tabs = self._area_scroll(col_izq)

        panel = ttk.Frame(marco)
        panel.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(12, 0))

        estado = {"activa": None, "paginas": {}, "botones": {}, "refrescos": {},
                  "entidades": list(entidades)}

        def prefijo(entidad):
            fondos = self._fondos_de_entidad(entidad)
            # El número de la izquierda es cuántos fondos están marcados
            marcados = contador(fondos) if contador else 0
            return f"[{marcados}/{len(fondos)}]  "

        def actualizar_titulos():
            for entidad, boton in estado["botones"].items():
                if boton.winfo_exists():
                    # El contador queda fijo; solo el nombre de la pestaña activa se mueve
                    self._registrar_marquee(boton, prefijo(entidad), entidad,
                                            VISIBLES_PESTANA,
                                            animar=(entidad == estado["activa"]))

        estado["actualizar_titulos"] = actualizar_titulos
        estado["mostrar"] = lambda entidad: mostrar(entidad)
        estado["ajustar_scroll"] = canvas_tabs._ajustar_scroll

        def mostrar(entidad):
            for otra, pagina in estado["paginas"].items():
                pagina.pack_forget()
                estado["botones"][otra].configure(style="Tab.TButton")
            estado["paginas"][entidad].pack(fill=tk.BOTH, expand=True)
            estado["botones"][entidad].configure(style="TabSel.TButton")
            estado["activa"] = entidad
            # El desplazamiento del nombre pasa a la pestaña que quedó activa
            actualizar_titulos()

        for entidad in entidades:
            pagina = ttk.Frame(panel)
            estado["paginas"][entidad] = pagina
            estado["refrescos"][entidad] = constructor(pagina, entidad)
            boton = ttk.Button(lista_tabs, style="Tab.TButton",
                               command=lambda e=entidad: mostrar(e))
            boton.pack(fill=tk.X, padx=4, pady=2)
            estado["botones"][entidad] = boton
            self._registrar_marquee(boton, prefijo(entidad), entidad, VISIBLES_PESTANA)

        # Al cerrar el diálogo se detiene la animación. La ventana se resuelve AHORA:
        # hacerlo dentro del callback consultaría un widget ya destruido.
        ventana = contenedor.winfo_toplevel()

        def _detener(evento, ventana=ventana):
            if evento.widget is ventana:
                self._marquee.clear()
                self._marquee_activo = False

        ventana.bind("<Destroy>", _detener, add="+")

        self._habilitar_rueda(canvas_tabs, lista_tabs)
        if entidades:
            mostrar(entidades[0])
        return estado

    def _habilitar_rueda(self, canvas, interior):
        """La rueda del mouse desplaza el canvas estando sobre cualquier punto del área."""
        def _rueda(event):
            if not canvas.winfo_exists():
                return "break"
            if not _desborda(canvas):
                canvas.yview_moveto(0)      # cabe entero: no hay nada que desplazar
                return "break"
            canvas.yview_scroll(-1 if event.delta > 0 else 1, "units")
            return "break"

        def _aplicar(widget):
            widget.bind("<MouseWheel>", _rueda)
            for hijo in widget.winfo_children():
                _aplicar(hijo)

        canvas.bind("<MouseWheel>", _rueda)
        _aplicar(interior)

    def _centrar_toplevel(self, ventana, ancho, alto):
        """Centra un popup en la pantalla, acotado al tamaño disponible."""
        self.root.update_idletasks()
        ancho = min(ancho, ventana.winfo_screenwidth() - 40)
        alto = min(alto, ventana.winfo_screenheight() - 80)
        x = (ventana.winfo_screenwidth() - ancho) // 2
        y = (ventana.winfo_screenheight() - alto) // 2
        ventana.geometry(f"{ancho}x{alto}+{x}+{y}")

    def _reposicionar_filtros_columna(self, event=None):
        """Alinea cada caja de filtro con el ancho de su columna."""
        x = 0
        for col_id in COL_IDS:
            ancho = self.tree.column(col_id, "width")
            self.filtro_entries[col_id].place(x=x, y=1, width=ancho, height=ALTO_FILTRO)
            x += ancho

    def _ordenar_por_columna(self, col_id):
        if self._sort_col == col_id:
            self._sort_asc = not self._sort_asc
        else:
            self._sort_col = col_id
            self._sort_asc = True
        for c in COL_IDS:
            flecha = ("  ▲" if self._sort_asc else "  ▼") if c == self._sort_col else ""
            self.tree.heading(c, text=COL_ETIQUETA[c] + flecha)
        self._aplicar_filtros()

    def _identificar_celda(self, event):
        fila = self.tree.identify_row(event.y)
        col = self.tree.identify_column(event.x)
        if not fila or not col:
            return None, None
        try:
            idx = int(col.lstrip("#")) - 1
        except ValueError:
            return None, None
        if idx < 0 or idx >= len(COL_IDS):
            return None, None
        return fila, COL_IDS[idx]

    def _aplicar_seleccion(self, fila_hasta, col_hasta):
        if self._sel_ancla is None:
            return
        fila_ini, col_ini = self._sel_ancla
        filas = list(self.tree.get_children())
        if fila_ini not in filas or fila_hasta not in filas:
            return
        i0, i1 = sorted((filas.index(fila_ini), filas.index(fila_hasta)))
        self._sel_filas = filas[i0:i1 + 1]
        c0, c1 = sorted((COL_IDS.index(col_ini), COL_IDS.index(col_hasta)))
        self._sel_cols = COL_IDS[c0:c1 + 1]
        self.tree.selection_set(self._sel_filas)
        self._actualizar_label_seleccion()

    def _actualizar_label_seleccion(self):
        n_filas = len(self._sel_filas)
        n_cols = len(self._sel_cols)
        if n_filas and n_cols:
            cols = ", ".join(COL_ETIQUETA[c] for c in self._sel_cols)
            self.label_seleccion.config(text=f"Selección: {n_filas} fila(s) × {n_cols} col. ({cols})")
        else:
            self.label_seleccion.config(text="")

    def _mostrar_menu_contextual(self, event):
        fila, col = self._identificar_celda(event)
        if fila and (fila not in self._sel_filas or col not in self._sel_cols):
            self._sel_ancla = (fila, col)
            self._aplicar_seleccion(fila, col)
        if self._sel_filas:
            self.menu_contextual.tk_popup(event.x_root, event.y_root)

    def _copiar_seleccion(self, event=None):
        if not self._sel_filas or not self._sel_cols:
            return
        idx_cols = [COL_IDS.index(c) for c in self._sel_cols]
        lineas = []
        for iid in self._sel_filas:
            valores = self.tree.item(iid, "values")
            lineas.append("\t".join(str(valores[i]) for i in idx_cols))
        self.root.clipboard_clear()
        self.root.clipboard_append("\n".join(lineas))

    def _on_tree_press(self, event):
        if self.tree.identify_region(event.x, event.y) != "cell":
            return None
        fila, col = self._identificar_celda(event)
        if not fila:
            return None
        self.tree.focus_set()
        self._sel_ancla = (fila, col)
        self._aplicar_seleccion(fila, col)
        return "break"

    def _on_tree_drag(self, event):
        if self._sel_ancla is None:
            return None
        fila, col = self._identificar_celda(event)
        if not fila:
            return None
        self._aplicar_seleccion(fila, col)
        return "break"

    def _on_tree_release(self, event):
        self._reposicionar_filtros_columna()

    def descargar_datos(self):
        threading.Thread(target=self._descargar_thread, daemon=True).start()

    def _descargar_thread(self):
        driver = None
        try:
            for f in glob.glob(os.path.join(self.download_dir, "*")):
                os.remove(f)

            self.label_status.config(text="Abriendo navegador...", foreground="#5B9BD5")
            self.root.update()

            url = "https://www.cmfchile.cl/institucional/estadisticas/fondos_cartola_diaria.php"

            options = webdriver.ChromeOptions()
            options.add_experimental_option("prefs", {
                "download.default_directory": self.download_dir,
                "download.prompt_for_download": False,
            })
            options.add_argument("--window-size=780,720")
            options.add_argument("--app=" + url)     # ventana limpia, sin barras del navegador
            driver = webdriver.Chrome(options=options)
            if not driver.current_url.startswith("http"):
                driver.get(url)

            # Primero las fechas: es lo que el usuario espera ver de inmediato
            desde = datetime.strptime(self.entry_desde.get(), "%Y-%m-%d").strftime("%d/%m/%Y")
            hasta = datetime.strptime(self.entry_hasta.get(), "%Y-%m-%d").strftime("%d/%m/%Y")
            espera = WebDriverWait(driver, 30)
            campo_desde = espera.until(EC.presence_of_element_located((By.NAME, "txt_inicio")))
            campo_hasta = driver.find_element(By.NAME, "txt_termino")
            campo_desde.clear()
            campo_desde.send_keys(desde)
            campo_hasta.clear()
            campo_hasta.send_keys(hasta)

            # Dejar a la vista solo el captcha y el botón, con el cursor ya puesto ahí
            try:
                driver.execute_script(JS_SOLO_FORMULARIO)
                # Traer la ventana al frente para que lo escrito vaya al captcha
                driver.switch_to.window(driver.current_window_handle)
                driver.find_element(By.NAME, "captcha").click()
            except Exception:
                pass                                 # si el sitio cambia, se sigue igual

            self.label_status.config(text="Resuelva el captcha y presione 'GENERAR ARCHIVO' en el navegador...", foreground="#5B9BD5")
            self.root.update()

            # El catálogo de nombres se lee después: no debe retrasar la pantalla
            self._guardar_catalogo_fondos(_extraer_catalogo_fondos(driver))

            archivo = self._esperar_descarga(timeout=300)
            driver.quit()
            driver = None

            self.label_status.config(text="Procesando archivo...", foreground="#5B9BD5")
            self.root.update()

            # Parsear CSV (la CMF usa ';' como separador y '.' como decimal)
            self.df = _normalizar_numericas(pd.read_csv(archivo, sep=";", encoding="latin-1"))
            self.df = _agregar_nombre_fondo(self.df, self._cargar_catalogo_fondos())

            try:
                os.remove(archivo)
            except OSError:
                pass

            self._poblar_entidades(sorted(str(e) for e in self.df.get("NOM_ADM", pd.Series(dtype=str)).dropna().unique()))
            self._poblar_fondos(self.df)

            for col_id, var in self.filtros_columna.items():
                var.set("")
                self.filtro_entries[col_id]._mostrar_placeholder()

            # Ajustar el rango de fechas visible a lo que realmente trae el archivo
            if "FECHA_INF" in self.df.columns:
                fechas = pd.to_datetime(self.df["FECHA_INF"].astype(str), format="%Y%m%d", errors="coerce")
                if fechas.notna().any():
                    self._date_desde_filter.set_date(fechas.min().date())
                    self._date_hasta_filter.set_date(fechas.max().date())

            self._aplicar_filtros()

            self.label_status.config(text=f"✓ {len(self.df)} registros cargados", foreground="#4CAF50")

        except Exception as e:
            self.label_status.config(text=f"✗ Error: {str(e)}", foreground="#F44336")
            messagebox.showerror("Error", str(e))
        finally:
            if driver:
                driver.quit()

    def _esperar_descarga(self, timeout):
        inicio = time.time()
        while time.time() - inicio < timeout:
            archivos = [f for f in glob.glob(os.path.join(self.download_dir, "*"))
                        if not f.endswith(".crdownload")]
            if archivos:
                time.sleep(1)  # asegura que Chrome terminó de escribir el archivo
                return archivos[0]
            time.sleep(1)
        raise Exception("Tiempo de espera agotado esperando el archivo (¿resolvió el captcha y presionó 'Generar Archivo'?)")

    def _poblar_entidades(self, nombres):
        preferencias = self._cargar_preferencias()
        self.entidades_disponibles = nombres

        # Si hay preferencias guardadas, usar solo las que están ahí; si no, todas
        if preferencias is not None and preferencias:
            default_checked = {nombre: (nombre in preferencias) for nombre in nombres}
        else:
            default_checked = {nombre: True for nombre in nombres}

        self.entidad_vars = {
            nombre: tk.BooleanVar(value=default_checked[nombre])
            for nombre in nombres
        }
        self._actualizar_texto_boton_entidad()

    # ------------------------------------------------------- indicadores del día
    def _pintar_indicadores(self, datos, dia):
        """Escribe los valores y sus fechas. Se llama siempre en el hilo de la UI."""
        if not self.label_indicadores.winfo_exists():
            return

        def pintar(texto, fechas, fondo):
            for w in (self.frame_indicadores, self.label_indicadores, self.label_ind_fechas):
                w.config(bg=fondo)
            self.label_indicadores.config(text=texto)
            self.label_ind_fechas.config(text=fechas)

        if not datos:
            guardados = self._cargar_prefs_json().get("indicadores") or {}
            if guardados.get("valores"):
                # Sin conexión: se muestra lo último conocido, dejando claro de cuándo es
                datos = guardados["valores"]
                pintar(self._texto_indicadores(datos),
                       "Sin conexión — " + self._texto_fechas(datos, guardados.get("dia")),
                       ACCENT_DARK)
            else:
                pintar("UF · UTM · Dólar: sin conexión", "", ACCENT_DARK)
            return

        pintar(self._texto_indicadores(datos), self._texto_fechas(datos, dia), BOTON_BG)
        self._indicadores_dia = dia
        prefs = self._cargar_prefs_json()
        prefs["indicadores"] = {"dia": dia, "valores": datos}
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(prefs, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

    @staticmethod
    def _texto_indicadores(datos):
        partes = []
        for clave, etiqueta, decimales in INDICADORES:
            if datos.get(clave) is not None:
                partes.append(f"{etiqueta} ${_formato_clp(datos[clave], decimales)}")
        return "   ·   ".join(partes) if partes else "UF · UTM · Dólar: sin datos"

    @staticmethod
    def _texto_fechas(datos, dia_consulta=None):
        """Fecha a la que corresponde cada valor.

        La UTM es mensual, así que su fecha es el día 1: cuando difiere de las demás
        se indica por separado en vez de dar una sola fecha para todo.
        """
        fechas = datos.get("fechas") or {}
        propias = {c: fechas.get(c) for c, _e, _d in INDICADORES if fechas.get(c)}
        if not propias:
            return f"Consultado el {_fecha_corta(dia_consulta)}" if dia_consulta else ""

        distintas = set(propias.values())
        if len(distintas) == 1:
            return f"Valores al {_fecha_corta(distintas.pop())}"

        etiquetas = {c: e for c, e, _d in INDICADORES}
        detalle = " · ".join(f"{etiquetas[c]} {_fecha_corta(f)}" for c, f in propias.items())
        return f"Valores al: {detalle}"

    def _refrescar_indicadores(self, forzar=False):
        """Consulta los indicadores en segundo plano; usa el caché si ya son de hoy."""
        hoy = datetime.now().strftime("%Y-%m-%d")
        if not forzar:
            guardados = self._cargar_prefs_json().get("indicadores") or {}
            if guardados.get("dia") == hoy and guardados.get("valores"):
                self._pintar_indicadores(guardados["valores"], hoy)
                return

        self.label_indicadores.config(text="UF · UTM · Dólar: actualizando…")
        self._pendiente_indicadores = None

        def trabajo():
            # Solo se deja el resultado en un atributo: Tkinter NO se puede tocar
            # desde otro hilo (root.after aquí lanza "main thread is not in main loop")
            self._pendiente_indicadores = (_consultar_indicadores(), hoy)

        threading.Thread(target=trabajo, daemon=True).start()
        self._esperar_indicadores()

    def _esperar_indicadores(self, intentos=0):
        """Recoge el resultado del hilo, siempre desde el hilo de la interfaz."""
        pendiente = self._pendiente_indicadores
        if pendiente is not None:
            self._pendiente_indicadores = None
            self._pintar_indicadores(*pendiente)
            return
        if intentos >= 60:                     # ~15 s: se da por fallida
            self._pintar_indicadores({}, datetime.now().strftime("%Y-%m-%d"))
            return
        try:
            self.root.after(250, lambda: self._esperar_indicadores(intentos + 1))
        except tk.TclError:
            pass                               # la ventana se cerró

    def _revisar_cambio_de_dia(self):
        """Si cambió la fecha con la app abierta, vuelve a pedir los indicadores."""
        hoy = datetime.now().strftime("%Y-%m-%d")
        if getattr(self, "_indicadores_dia", None) != hoy:
            self._refrescar_indicadores()
        try:
            self.root.after(REVISION_INDICADORES_MS, self._revisar_cambio_de_dia)
        except tk.TclError:
            pass                     # la ventana se cerró

    def _abrir_informacion(self):
        """Ficha de la aplicación, con el logo de fondo como marca de agua."""
        ANCHO, ALTO = 680, 470
        info = tk.Toplevel(self.root)
        info.title("Información")
        info.config(bg=BG_PRIMARY)
        info.transient(self.root)
        self._centrar_toplevel(info, ANCHO, ALTO)

        # Todo se dibuja en un lienzo: así el logo de fondo queda visible detrás
        # del texto (un Frame normal lo taparía, porque Tk no tiene transparencia)
        lienzo = tk.Canvas(info, bg=BG_PRIMARY, highlightthickness=0)
        lienzo.pack(fill=tk.BOTH, expand=True)

        self._fondo_info = _cargar_logo(420, atenuar=0.14, fondo=BG_PRIMARY)
        if self._fondo_info is not None:
            lienzo.create_image(ANCHO // 2, ALTO // 2 + 20, image=self._fondo_info)

        self._logo_info = _cargar_logo(74)
        if self._logo_info is not None:
            lienzo.create_image(28, 26, image=self._logo_info, anchor="nw")

        lienzo.create_text(118, 34, text=APP_NOMBRE, anchor="nw", width=530,
                           font=("Segoe UI", 17, "bold"), fill=TEXT_PRIMARY)
        lienzo.create_text(118, 66, text=f"Versión {APP_VERSION}", anchor="nw",
                           font=FONT_NORMAL, fill=TEXT_SECONDARY)

        def fila(y, titulo, texto, url=None):
            lienzo.create_text(30, y, text=titulo, anchor="nw", font=FONT_BOLD,
                               fill=TEXT_PRIMARY)
            if url:
                marca = lienzo.create_text(
                    215, y, text=texto, anchor="nw", width=430,
                    font=("Segoe UI", 13, "underline"), fill="#1A5FA8")
                lienzo.tag_bind(marca, "<Button-1>", lambda _e, u=url: webbrowser.open(u))
                lienzo.tag_bind(marca, "<Enter>",
                                lambda _e: lienzo.config(cursor="hand2"))
                lienzo.tag_bind(marca, "<Leave>", lambda _e: lienzo.config(cursor=""))
            else:
                lienzo.create_text(215, y, text=texto, anchor="nw", width=430,
                                   font=FONT_NORMAL, fill=TEXT_PRIMARY)

        fila(140, "Qué hace:", "Descarga la cartola diaria de fondos mutuos publicada "
                               "por la CMF y permite filtrarla y exportarla a Excel.")
        fila(232, "Fuente de datos:", "Cartola diaria de fondos mutuos — CMF Chile", URL_CMF)
        fila(288, "Desarrollado por:", APP_AUTOR)
        if APP_LINKEDIN:
            fila(340, "LinkedIn:", APP_LINKEDIN, APP_LINKEDIN)
        else:
            fila(340, "LinkedIn:", "(sin configurar)")

        cerrar = ttk.Button(lienzo, text="Cerrar", command=info.destroy)
        lienzo.create_window(ANCHO - 30, ALTO - 26, window=cerrar, anchor="se")

    def _abrir_preferencias(self):
        if not self.entidades_disponibles:
            messagebox.showinfo("Preferencias", "Descarga datos primero para configurar preferencias")
            return

        pref = tk.Toplevel(self.root)
        pref.title("Preferencias")
        pref.config(bg=BG_PRIMARY)
        pref.transient(self.root)
        self._centrar_toplevel(pref, 1240, 820)

        # Dos filas ancladas al fondo antes que nada más: así siempre se ven,
        # y en ventanas angostas Guardar/Cancelar no se salen por la derecha
        barra = ttk.Frame(pref)
        barra.pack(side=tk.BOTTOM, fill=tk.X, padx=14, pady=(4, 14))
        barra_pestana = ttk.Frame(pref)
        barra_pestana.pack(side=tk.BOTTOM, fill=tk.X, padx=14, pady=(6, 0))

        ttk.Label(pref, text="1) Entidades a mostrar por defecto",
                  font=FONT_BOLD).pack(anchor=tk.W, padx=14, pady=(14, 4))

        var_buscar_ent = tk.StringVar()
        entry_buscar_ent = ttk.Entry(pref, textvariable=var_buscar_ent, font=FONT_NORMAL)
        entry_buscar_ent.pack(fill=tk.X, padx=14, pady=(0, 6))
        leer_buscar_ent = _poner_placeholder(entry_buscar_ent, var_buscar_ent,
                                             "escribe parte del nombre de la entidad...")

        # --- Entidades ---
        frame_ent = ttk.Frame(pref)
        frame_ent.pack(fill=tk.X, padx=14)
        canvas_ent, interior_ent = self._area_scroll(frame_ent, alto=170)

        prefs_ent = self._cargar_preferencias() or []
        pref_ent_vars, casillas_ent = {}, {}
        for nombre in self.entidades_disponibles:
            marcado = (nombre in prefs_ent) if prefs_ent else True
            pref_ent_vars[nombre] = tk.BooleanVar(value=marcado)
            casillas_ent[nombre] = ttk.Checkbutton(interior_ent, text=nombre,
                                                   variable=pref_ent_vars[nombre],
                                                   command=lambda: reconstruir())
            casillas_ent[nombre].pack(anchor=tk.W, padx=8, pady=2)
        self._habilitar_rueda(canvas_ent, interior_ent)

        def filtrar_entidades(*_):
            texto = leer_buscar_ent().strip().lower()
            for nombre in self.entidades_disponibles:
                casillas_ent[nombre].pack_forget()
            for nombre in self.entidades_disponibles:
                if texto in nombre.lower():
                    casillas_ent[nombre].pack(anchor=tk.W, padx=8, pady=2)

        var_buscar_ent.trace_add("write", filtrar_entidades)

        fila_ent = ttk.Frame(pref)
        fila_ent.pack(fill=tk.X, padx=14, pady=(6, 10))
        ttk.Button(fila_ent, text="Marcar todas las entidades",
                   command=lambda: [v.set(True) for v in pref_ent_vars.values()] or reconstruir()
                   ).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(fila_ent, text="Desmarcar todas",
                   command=lambda: [v.set(False) for v in pref_ent_vars.values()] or reconstruir()
                   ).pack(side=tk.LEFT)

        ttk.Label(pref, text="2) Fondos a mostrar por defecto (una pestaña por entidad)",
                  font=FONT_BOLD).pack(anchor=tk.W, padx=14, pady=(4, 4))

        var_buscar_fondo = tk.StringVar()
        entry_buscar_fondo = ttk.Entry(pref, textvariable=var_buscar_fondo, font=FONT_NORMAL)
        entry_buscar_fondo.pack(fill=tk.X, padx=14, pady=(0, 6))
        leer_buscar_fondo = _poner_placeholder(entry_buscar_fondo, var_buscar_fondo,
                                               "escribe el número o el nombre del fondo...")

        # --- Fondos por entidad, en pestañas ---
        contenedor_nb = ttk.Frame(pref)
        contenedor_nb.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 8))

        # Copias editables: no se toca el estado real hasta presionar Guardar
        pref_fondo_vars = {f: tk.BooleanVar(value=self.fondo_vars[f].get())
                           for f in self.fondos_disponibles}
        pref_series = {f: set(self._series_marcadas(f)) for f in self.fondos_disponibles}
        estado = {"marco": None, "pestanas": None, "entidades": []}

        def reconstruir():
            if estado["marco"] is not None:
                estado["marco"].destroy()
            marcadas = [n for n in self.entidades_disponibles if pref_ent_vars[n].get()]
            con_fondos = {self.fondo_entidad.get(f, "") for f in self.fondos_disponibles}
            estado["entidades"] = [n for n in marcadas if n in con_fondos]

            marco = ttk.Frame(contenedor_nb)
            marco.pack(fill=tk.BOTH, expand=True)
            estado["marco"] = marco

            if not estado["entidades"]:
                ttk.Label(marco, text="Marca al menos una entidad arriba",
                          font=FONT_NORMAL).pack(padx=20, pady=20)
                estado["pestanas"] = None
                return

            estado["pestanas"] = self._pestanas_por_entidad(
                marco, estado["entidades"],
                lambda pagina, entidad: self._llenar_pestana_prefs(
                    pagina, self._fondos_de_entidad(entidad), pref_fondo_vars, pref_series,
                    refrescar_titulos=lambda: estado["pestanas"]["actualizar_titulos"]()),
                contador=lambda fondos: sum(1 for f in fondos if pref_fondo_vars[f].get()))
            filtrar_fondos()

        def filtrar_fondos(*_):
            if not estado["pestanas"]:
                return
            texto = leer_buscar_fondo().strip().lower()
            for refrescar in estado["pestanas"]["refrescos"].values():
                refrescar(texto)

        var_buscar_fondo.trace_add("write", filtrar_fondos)
        reconstruir()

        def entidad_activa():
            return estado["pestanas"]["activa"] if estado["pestanas"] else None

        def marcar_pestana(valor):
            entidad = entidad_activa()
            if entidad is None:
                return
            for f in self._fondos_de_entidad(entidad):
                pref_fondo_vars[f].set(valor)

        def guardar():
            seleccionadas = [n for n, v in pref_ent_vars.items() if v.get()]
            for nombre, var in self.entidad_vars.items():
                var.set(nombre in seleccionadas)
            for f, var in self.fondo_vars.items():
                # Solo cuentan los fondos de entidades marcadas
                var.set(pref_fondo_vars[f].get() and
                        self.fondo_entidad.get(f, "") in seleccionadas)
            for f, elegidas in pref_series.items():
                self.fondo_series[f] = set(elegidas)
            self._guardar_preferencias()
            self._actualizar_texto_boton_entidad()
            self._actualizar_texto_boton_fondo()
            self._aplicar_filtros()
            messagebox.showinfo("Éxito", "Preferencias guardadas")
            pref.destroy()

        ttk.Button(barra_pestana, text="Marcar esta pestaña",
                   command=lambda: marcar_pestana(True)).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(barra_pestana, text="Desmarcar esta pestaña",
                   command=lambda: marcar_pestana(False)).pack(side=tk.LEFT)
        ttk.Button(barra, text="Cancelar", command=pref.destroy).pack(side=tk.RIGHT)
        ttk.Button(barra, text="Guardar", command=guardar).pack(side=tk.RIGHT, padx=(0, 8))

    def _llenar_pestana_prefs(self, pestana, fondos, variables, series_editables,
                              refrescar_titulos=None):
        """Pestaña de preferencias: número, nombre y serie por defecto de cada fondo."""
        ANCHO_NUM, ANCHO_SERIE = 130, 200

        cabecera = ttk.Frame(pestana)
        cabecera.pack(fill=tk.X, padx=(8, 24), pady=(0, 4))
        ttk.Label(cabecera, text="Número", font=FONT_BOLD).grid(row=0, column=0, sticky="w")
        ttk.Label(cabecera, text="Nombre Fondo", font=FONT_BOLD).grid(row=0, column=1, sticky="w")
        ttk.Label(cabecera, text="Serie", font=FONT_BOLD).grid(row=0, column=2, sticky="e")
        cabecera.columnconfigure(0, minsize=ANCHO_NUM)
        cabecera.columnconfigure(1, weight=1)
        cabecera.columnconfigure(2, minsize=ANCHO_SERIE)

        canvas, interior = self._area_scroll(pestana)
        interior.columnconfigure(0, minsize=ANCHO_NUM)
        interior.columnconfigure(1, weight=1)
        interior.columnconfigure(2, minsize=ANCHO_SERIE)

        widgets, busqueda = {}, {}
        for indice, f in enumerate(fondos):
            nombre = self.fondo_nombre.get(f, "")
            busqueda[f] = f"{f} {nombre}".lower()
            chk = ttk.Checkbutton(interior, text=str(f), variable=variables[f],
                                  command=refrescar_titulos)
            chk.grid(row=indice, column=0, sticky="w", padx=8, pady=2)
            etiqueta = ttk.Label(interior, text=nombre or "—")
            etiqueta.grid(row=indice, column=1, sticky="w", padx=(4, 8), pady=2)
            boton = ttk.Button(interior, width=18,
                               text=self._resumen_series(f, series_editables.get(f)))
            boton.configure(command=lambda run=f, b=boton: self._elegir_series_en_prefs(
                run, series_editables, b))
            boton.grid(row=indice, column=2, sticky="e", padx=(14, 8), pady=2)
            widgets[f] = (chk, etiqueta, boton)

        self._habilitar_rueda(canvas, interior)

        def refrescar(texto=""):
            fila = 0
            for f in fondos:
                chk, etiqueta, boton = widgets[f]
                if texto in busqueda[f]:
                    chk.grid(row=fila, column=0, sticky="w", padx=8, pady=2)
                    etiqueta.grid(row=fila, column=1, sticky="w", padx=(4, 8), pady=2)
                    boton.grid(row=fila, column=2, sticky="e", padx=(14, 8), pady=2)
                    fila += 1
                else:
                    chk.grid_remove()
                    etiqueta.grid_remove()
                    boton.grid_remove()
            canvas._ajustar_scroll()      # el área desplazable se encoge con la lista

        return refrescar

    def _elegir_series_en_prefs(self, run, series_editables, boton):
        """Como el popup de series, pero escribe en la copia editable de Preferencias."""
        elegidas = self._abrir_series_fondo(run, seleccion_inicial=series_editables.get(run))
        if elegidas is not None:
            series_editables[run] = elegidas
            if boton.winfo_exists():
                boton.configure(text=self._resumen_series(run, elegidas))

    def _abrir_selector_entidades(self):
        if self.popup_entidades is not None and self.popup_entidades.winfo_exists():
            self.popup_entidades.destroy()
            self.popup_entidades = None
            return

        popup = tk.Toplevel(self.root)
        popup.title("Seleccionar entidades")
        popup.config(bg=BG_PRIMARY)
        popup.transient(self.root)
        self._centrar_toplevel(popup, 840, 660)
        self.popup_entidades = popup

        if not self.entidades_disponibles:
            ttk.Label(popup, text="Descarga datos primero", font=FONT_NORMAL).pack(padx=20, pady=20)
            ttk.Button(popup, text="Cerrar", command=popup.destroy).pack(pady=8)
            return

        ttk.Label(popup, text="Buscar:").pack(anchor=tk.W, padx=10, pady=(10, 0))
        var_busqueda_entidad = tk.StringVar()
        entry_ent = ttk.Entry(popup, textvariable=var_busqueda_entidad, font=FONT_NORMAL)
        entry_ent.pack(fill=tk.X, padx=10, pady=(2, 5))
        leer_ent = _poner_placeholder(entry_ent, var_busqueda_entidad,
                                      "escribe parte del nombre de la entidad...")

        frame_medio = ttk.Frame(popup)
        frame_medio.pack(fill=tk.BOTH, expand=True, padx=10, pady=(0, 10))

        # Los botones reservan su ancho antes que la lista expandible
        frame_botones = ttk.Frame(frame_medio)
        frame_botones.pack(side=tk.RIGHT, fill=tk.Y, padx=(10, 0))

        frame_lista = ttk.Frame(frame_medio)
        frame_lista.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        canvas, frame_checks = self._area_scroll(frame_lista)

        checkbuttons_entidad = {}
        for nombre in self.entidades_disponibles:
            chk = ttk.Checkbutton(frame_checks, text=_acortar(nombre, 62),
                                   variable=self.entidad_vars[nombre],
                                   command=self._on_entidad_toggle)
            chk.pack(anchor=tk.W, padx=8, pady=3)
            checkbuttons_entidad[nombre] = chk

        self._habilitar_rueda(canvas, frame_checks)

        def filtrar_lista_entidades(*_):
            texto = leer_ent().strip().lower()
            for nombre in self.entidades_disponibles:
                checkbuttons_entidad[nombre].pack_forget()
            for nombre in self.entidades_disponibles:
                if texto in nombre.lower():
                    checkbuttons_entidad[nombre].pack(anchor=tk.W, padx=8, pady=3)
            canvas._ajustar_scroll()

        var_busqueda_entidad.trace_add("write", filtrar_lista_entidades)

        def seleccionar_todo():
            for v in self.entidad_vars.values():
                v.set(True)
            self._on_entidad_toggle()

        def deseleccionar_todo():
            for v in self.entidad_vars.values():
                v.set(False)
            self._on_entidad_toggle()

        # Columna de botones al costado del listado, uno abajo de otro
        ttk.Button(frame_botones, text="Seleccionar todo", width=18,
                   command=seleccionar_todo).pack(side=tk.TOP, pady=4)
        ttk.Button(frame_botones, text="Deseleccionar todo", width=18,
                   command=deseleccionar_todo).pack(side=tk.TOP, pady=4)
        ttk.Button(frame_botones, text="Cerrar", width=18, command=popup.destroy).pack(side=tk.TOP, pady=(20, 4))

    def _on_entidad_toggle(self):
        self._guardar_preferencias()
        self._actualizar_texto_boton_entidad()
        self._actualizar_texto_boton_fondo()
        self._aplicar_filtros()

    def _actualizar_texto_boton_entidad(self):
        total = len(self.entidad_vars)
        seleccionadas = sum(1 for v in self.entidad_vars.values() if v.get())
        if total == 0:
            texto = "Entidad ▾"
        elif seleccionadas == total:
            texto = "Entidad: Todas ▾"
        elif seleccionadas == 0:
            texto = "Entidad: Ninguna ▾"
        else:
            texto = f"Entidad: {seleccionadas} seleccionadas ▾"
        self.btn_entidad.config(text=texto)

    def _poblar_fondos(self, df):
        """Registra cada fondo con su entidad, su nombre y las series que tiene."""
        self.fondo_entidad = {}
        self.fondo_nombre = {}
        series_tmp = {}

        if "RUN_FM" in df.columns:
            runs = df["RUN_FM"].astype(str).tolist()
            entidades = df["NOM_ADM"].astype(str).tolist() if "NOM_ADM" in df.columns else [""] * len(runs)
            nombres = df["NOMBRE_FM"].astype(str).tolist() if "NOMBRE_FM" in df.columns else [""] * len(runs)
            series = df["SERIE"].astype(str).tolist() if "SERIE" in df.columns else [""] * len(runs)
            for run, ent, nom, ser in zip(runs, entidades, nombres, series):
                self.fondo_entidad.setdefault(run, ent)
                if nom and nom != "nan" and run not in self.fondo_nombre:
                    self.fondo_nombre[run] = nom
                if ser:
                    series_tmp.setdefault(run, set()).add(ser)

        self.series_por_fondo = {r: sorted(v) for r, v in series_tmp.items()}
        self.fondos_disponibles = sorted(self.fondo_entidad.keys(), key=_clave_fondo)

        # Sin preferencias previas los fondos arrancan DESMARCADOS: se eligen a mano
        guardados = self._cargar_preferencias_fondos()
        marcado = ({f: (f in guardados) for f in self.fondos_disponibles} if guardados
                   else {f: False for f in self.fondos_disponibles})
        self.fondo_vars = {f: tk.BooleanVar(value=marcado[f]) for f in self.fondos_disponibles}

        # Series marcadas por fondo; por defecto todas. Se descartan las que ya no existen
        series_guardadas = self._cargar_preferencias_series_fondo() or {}
        self.fondo_series = {}
        for f in self.fondos_disponibles:
            todas = self.series_por_fondo.get(f, [])
            guardadas = series_guardadas.get(f)
            if isinstance(guardadas, str):        # formato antiguo: una sola serie
                guardadas = [guardadas]
            validas = [s for s in (guardadas or []) if s in todas]
            self.fondo_series[f] = set(validas) if validas else set(todas)

        self._actualizar_texto_boton_fondo()

    def _series_marcadas(self, run):
        return self.fondo_series.get(run, set(self.series_por_fondo.get(run, [])))

    def _series_acotadas(self, run):
        """True si el fondo muestra solo una parte de sus series."""
        todas = set(self.series_por_fondo.get(run, []))
        return bool(todas) and self._series_marcadas(run) != todas

    def _resumen_series(self, run, conjunto=None):
        todas = self.series_por_fondo.get(run, [])
        if not todas:
            return "Sin series"
        marcadas = self._series_marcadas(run) if conjunto is None else conjunto
        if len(marcadas) == len(todas):
            return f"Todas ({len(todas)})"
        if not marcadas:
            return "Ninguna serie"
        if len(marcadas) <= 3:
            return ", ".join(sorted(marcadas))
        return f"{len(marcadas)} de {len(todas)}"

    def _etiqueta_fondo(self, run):
        nombre = self.fondo_nombre.get(run, "")
        return f"{run} - {nombre}" if nombre else str(run)

    def _fondos_de_entidades_seleccionadas(self):
        entidades = {n for n, v in self.entidad_vars.items() if v.get()}
        return [f for f in self.fondos_disponibles if self.fondo_entidad.get(f) in entidades]

    def _abrir_selector_fondos(self):
        if self.popup_fondos is not None and self.popup_fondos.winfo_exists():
            self.popup_fondos.destroy()
            self.popup_fondos = None
            return

        popup = tk.Toplevel(self.root)
        popup.title("Seleccionar fondos y series")
        popup.config(bg=BG_PRIMARY)
        popup.transient(self.root)
        self._centrar_toplevel(popup, 1240, 780)
        self.popup_fondos = popup

        entidades = self._entidades_con_fondos()
        if not entidades:
            mensaje = ("Descarga datos primero" if not self.fondos_disponibles
                       else "No hay fondos: selecciona al menos una entidad")
            ttk.Label(popup, text=mensaje, font=FONT_NORMAL).pack(padx=20, pady=20)
            ttk.Button(popup, text="Cerrar", command=popup.destroy).pack(pady=8)
            return

        ttk.Label(popup, text="Marca los fondos que quieres ver y elige la serie de cada uno:",
                  font=FONT_BOLD).pack(anchor=tk.W, padx=14, pady=(14, 6))
        # Dos buscadores independientes: uno acota las entidades (las pestañas de la
        # izquierda) y el otro los fondos que se listan dentro de cada pestaña
        frame_buscar = ttk.Frame(popup)
        frame_buscar.pack(fill=tk.X, padx=14, pady=(2, 8))
        frame_buscar.columnconfigure(0, weight=1, uniform="buscar")
        frame_buscar.columnconfigure(1, weight=1, uniform="buscar")

        ttk.Label(frame_buscar, text="Buscar entidad:").grid(row=0, column=0, sticky="w")
        var_buscar_ent = tk.StringVar()
        entry_buscar_ent = ttk.Entry(frame_buscar, textvariable=var_buscar_ent, font=FONT_NORMAL)
        entry_buscar_ent.grid(row=1, column=0, sticky="ew", padx=(0, 6))
        leer_buscar_ent = _poner_placeholder(entry_buscar_ent, var_buscar_ent,
                                             "nombre de la entidad...")

        ttk.Label(frame_buscar, text="Buscar fondo:").grid(row=0, column=1, sticky="w", padx=(6, 0))
        var_buscar_fondo = tk.StringVar()
        entry_buscar_fondo = ttk.Entry(frame_buscar, textvariable=var_buscar_fondo,
                                       font=FONT_NORMAL)
        entry_buscar_fondo.grid(row=1, column=1, sticky="ew", padx=(6, 0))
        leer_buscar_fondo = _poner_placeholder(entry_buscar_fondo, var_buscar_fondo,
                                               "número o nombre del fondo...")

        frame_medio = ttk.Frame(popup)
        frame_medio.pack(fill=tk.BOTH, expand=True, padx=14, pady=(0, 14))

        # La columna de botones se empaqueta ANTES del área expandible para que
        # reserve su ancho; si no, las pestañas se lo comen y los botones desaparecen
        frame_botones = ttk.Frame(frame_medio)
        frame_botones.pack(side=tk.RIGHT, fill=tk.Y, padx=(14, 0))

        contenedor_tabs = ttk.Frame(frame_medio)
        contenedor_tabs.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        # Una pestaña por entidad, con sus propios fondos
        widgets = {}
        pestanas = self._pestanas_por_entidad(
            contenedor_tabs, entidades,
            lambda pagina, entidad: self._llenar_pestana_fondos(
                pagina, self._fondos_de_entidad(entidad), widgets, con_series=True),
            contador=lambda fondos: sum(1 for f in fondos if self.fondo_vars[f].get()))

        # Que el contador de cada pestaña se actualice al marcar/desmarcar.
        # Se suelta al cerrar el diálogo para no tocar widgets ya destruidos.
        self._refrescar_tabs_fondos = pestanas["actualizar_titulos"]

        def _soltar(evento):
            if evento.widget is popup:
                self._refrescar_tabs_fondos = None
                self.popup_fondos = None

        popup.bind("<Destroy>", _soltar, add="+")

        def filtrar_fondos(*_):
            """El buscador de fondos acota lo que se lista dentro de cada pestaña."""
            texto = leer_buscar_fondo().strip().lower()
            for refrescar in pestanas["refrescos"].values():
                refrescar(texto)

        def filtrar_entidades(*_):
            """El buscador de entidades muestra u oculta las pestañas de la izquierda."""
            texto = leer_buscar_ent().strip().lower()
            for entidad, boton in pestanas["botones"].items():
                if not boton.winfo_exists():
                    continue
                if texto in entidad.lower():
                    boton.pack(fill=tk.X, padx=4, pady=2)
                else:
                    boton.pack_forget()
            pestanas["ajustar_scroll"]()   # con menos pestañas, menos área desplazable
            # Si la pestaña abierta dejó de estar visible, se pasa a la primera que sí lo está
            visibles = [e for e in entidades if texto in e.lower()]
            if visibles and pestanas["activa"] not in visibles:
                pestanas["mostrar"](visibles[0])

        var_buscar_fondo.trace_add("write", filtrar_fondos)
        var_buscar_ent.trace_add("write", filtrar_entidades)

        def entidad_activa():
            return pestanas["activa"] or entidades[0]

        def marcar(valor):
            for f in self._fondos_de_entidad(entidad_activa()):
                self.fondo_vars[f].set(valor)
            self._on_fondo_toggle()

        def series_todas():
            for f in self._fondos_de_entidad(entidad_activa()):
                self.fondo_series[f] = set(self.series_por_fondo.get(f, []))
                boton = widgets.get(f, (None, None, None))[1]
                if boton is not None and boton.winfo_exists():
                    boton.configure(text=self._resumen_series(f))
            self._on_fondo_toggle()

        def marcar_todo_global(valor):
            """Marca o desmarca los fondos de TODAS las entidades de una vez."""
            for entidad in entidades:
                for f in self._fondos_de_entidad(entidad):
                    self.fondo_vars[f].set(valor)
            self._on_fondo_toggle()

        # Cerrar va anclado al pie de la columna y se empaqueta primero: así queda
        # visible aunque el resto de los botones no alcancen a caber
        ttk.Button(frame_botones, text="Cerrar", width=22,
                   command=popup.destroy).pack(side=tk.BOTTOM, pady=(14, 4))

        ttk.Label(frame_botones, text="Todas las entidades:",
                  font=FONT_BOLD).pack(side=tk.TOP, pady=(0, 4))
        ttk.Button(frame_botones, text="Marcar TODOS", width=22,
                   command=lambda: marcar_todo_global(True)).pack(side=tk.TOP, pady=3)
        ttk.Button(frame_botones, text="Desmarcar TODOS", width=22,
                   command=lambda: marcar_todo_global(False)).pack(side=tk.TOP, pady=3)

        ttk.Label(frame_botones, text="Solo esta pestaña:",
                  font=FONT_BOLD).pack(side=tk.TOP, pady=(16, 4))
        ttk.Button(frame_botones, text="Seleccionar todo", width=22,
                   command=lambda: marcar(True)).pack(side=tk.TOP, pady=3)
        ttk.Button(frame_botones, text="Deseleccionar todo", width=22,
                   command=lambda: marcar(False)).pack(side=tk.TOP, pady=3)
        ttk.Button(frame_botones, text="Series: todas", width=22,
                   command=series_todas).pack(side=tk.TOP, pady=3)

    def _entidades_con_fondos(self):
        """Entidades marcadas que efectivamente tienen fondos en los datos."""
        marcadas = [n for n, v in self.entidad_vars.items() if v.get()]
        con_fondos = {self.fondo_entidad.get(f, "") for f in self.fondos_disponibles}
        return [n for n in marcadas if n in con_fondos]

    def _fondos_de_entidad(self, entidad):
        return [f for f in self.fondos_disponibles if self.fondo_entidad.get(f) == entidad]

    def _llenar_pestana_fondos(self, pestana, fondos, widgets, con_series):
        """Dibuja los fondos en columnas; devuelve la función que aplica el buscador."""
        ANCHO_NUM, ANCHO_SERIE = 130, 200

        # Encabezado fijo, alineado con las columnas de abajo
        cabecera = ttk.Frame(pestana)
        cabecera.pack(fill=tk.X, padx=(8, 24), pady=(0, 4))
        ttk.Label(cabecera, text="Número", font=FONT_BOLD).grid(row=0, column=0, sticky="w")
        ttk.Label(cabecera, text="Nombre Fondo", font=FONT_BOLD).grid(row=0, column=1, sticky="w")
        if con_series:
            ttk.Label(cabecera, text="Serie", font=FONT_BOLD).grid(row=0, column=2, sticky="e")
        cabecera.columnconfigure(0, minsize=ANCHO_NUM)
        cabecera.columnconfigure(1, weight=1)
        cabecera.columnconfigure(2, minsize=ANCHO_SERIE)

        canvas, interior = self._area_scroll(pestana)
        interior.columnconfigure(0, minsize=ANCHO_NUM)
        interior.columnconfigure(1, weight=1)
        interior.columnconfigure(2, minsize=ANCHO_SERIE)

        busqueda = {}
        for indice, f in enumerate(fondos):
            nombre = self.fondo_nombre.get(f, "")
            busqueda[f] = f"{f} {nombre} {self.fondo_entidad.get(f, '')}".lower()
            chk = ttk.Checkbutton(interior, text=str(f), variable=self.fondo_vars[f],
                                  command=self._on_fondo_toggle)
            chk.grid(row=indice, column=0, sticky="w", padx=8, pady=3)
            # Nombre completo, en su propia columna
            etiqueta = ttk.Label(interior, text=nombre or "—")
            etiqueta.grid(row=indice, column=1, sticky="w", padx=(4, 8), pady=3)
            boton = None
            if con_series:
                boton = ttk.Button(interior, width=18, text=self._resumen_series(f))
                boton.configure(command=lambda run=f, b=boton: self._abrir_series_fondo(run, b))
                boton.grid(row=indice, column=2, sticky="e", padx=(14, 8), pady=3)
            widgets[f] = (chk, boton, etiqueta)

        self._habilitar_rueda(canvas, interior)

        def refrescar(texto):
            fila = 0
            for f in fondos:
                chk, boton, etiqueta = widgets[f]
                if texto in busqueda[f]:
                    chk.grid(row=fila, column=0, sticky="w", padx=8, pady=3)
                    etiqueta.grid(row=fila, column=1, sticky="w", padx=(4, 8), pady=3)
                    if boton is not None:
                        boton.grid(row=fila, column=2, sticky="e", padx=(14, 8), pady=3)
                    fila += 1
                else:
                    chk.grid_remove()
                    etiqueta.grid_remove()
                    if boton is not None:
                        boton.grid_remove()
            canvas._ajustar_scroll()      # el área desplazable se encoge con la lista

        return refrescar

    def _abrir_series_fondo(self, run, boton=None, seleccion_inicial=None):
        """Popup con una casilla por serie del fondo.

        Modo normal: aplica la elección a `fondo_series` y devuelve None.
        Con `seleccion_inicial`: no toca el estado global y devuelve el conjunto
        elegido (o None si se cancela), para que Preferencias lo confirme al Guardar.
        """
        modo_prefs = seleccion_inicial is not None
        todas = self.series_por_fondo.get(run, [])
        if not todas:
            messagebox.showinfo("Series", "Este fondo no tiene series en los datos cargados")
            return None

        popup = tk.Toplevel(self.root)
        popup.title(f"Series de {self._etiqueta_fondo(run)}")
        popup.config(bg=BG_PRIMARY)
        popup.transient(self.root)
        popup.grab_set()
        alto = min(760, 300 + 34 * len(todas))
        self._centrar_toplevel(popup, 720, alto)

        ttk.Label(popup, text=self._etiqueta_fondo(run), font=FONT_BOLD,
                  wraplength=520).pack(anchor=tk.W, padx=14, pady=(14, 2))
        ttk.Label(popup, text="Marca las series que quieres mostrar:").pack(
            anchor=tk.W, padx=14, pady=(0, 4))

        var_buscar = tk.StringVar()
        entry_buscar = ttk.Entry(popup, textvariable=var_buscar, font=FONT_NORMAL)
        entry_buscar.pack(fill=tk.X, padx=14, pady=(0, 8))
        leer_buscar = _poner_placeholder(entry_buscar, var_buscar, "buscar serie...")

        # Una sola fila anclada al fondo ANTES del área scrollable. Se usa grid con
        # cuatro columnas uniformes: los botones quedan del mismo ancho, alineados,
        # y se reparten el espacio disponible sin cortarse
        barra = ttk.Frame(popup)
        barra.pack(side=tk.BOTTOM, fill=tk.X, padx=14, pady=(10, 12))
        for col in range(4):
            barra.columnconfigure(col, weight=1, uniform="acciones")

        marco = ttk.Frame(popup)
        marco.pack(fill=tk.BOTH, expand=True, padx=14)
        canvas, interior = self._area_scroll(marco)

        marcadas = seleccion_inicial if modo_prefs else self._series_marcadas(run)
        variables = {s: tk.BooleanVar(value=(s in marcadas)) for s in todas}
        casillas = {}
        for s in todas:
            casillas[s] = ttk.Checkbutton(interior, text=s, variable=variables[s])
            casillas[s].pack(anchor=tk.W, padx=8, pady=3)
        self._habilitar_rueda(canvas, interior)

        def filtrar_series(*_):
            texto = leer_buscar().strip().lower()
            for s in todas:
                casillas[s].pack_forget()
            for s in todas:
                if texto in s.lower():
                    casillas[s].pack(anchor=tk.W, padx=8, pady=3)
            canvas._ajustar_scroll()

        var_buscar.trace_add("write", filtrar_series)

        resultado = {"valor": None}

        def aplicar():
            elegidas = {s for s, v in variables.items() if v.get()}
            if modo_prefs:
                # Preferencias confirma al Guardar: aquí solo se devuelve
                resultado["valor"] = elegidas
            else:
                self.fondo_series[run] = elegidas
                self._guardar_preferencias()
                self._actualizar_texto_boton_fondo()
                self._aplicar_filtros()
                if boton is not None and boton.winfo_exists():
                    boton.configure(text=self._resumen_series(run))
            popup.destroy()

        ttk.Button(barra, text="Marcar todas",
                   command=lambda: [v.set(True) for v in variables.values()]
                   ).grid(row=0, column=0, sticky="ew", padx=(0, 4))
        ttk.Button(barra, text="Desmarcar todas",
                   command=lambda: [v.set(False) for v in variables.values()]
                   ).grid(row=0, column=1, sticky="ew", padx=4)
        ttk.Button(barra, text="Aplicar", command=aplicar
                   ).grid(row=0, column=2, sticky="ew", padx=4)
        ttk.Button(barra, text="Cancelar", command=popup.destroy
                   ).grid(row=0, column=3, sticky="ew", padx=(4, 0))

        self.root.wait_window(popup)
        return resultado["valor"]

    def _on_fondo_toggle(self):
        self._guardar_preferencias()
        self._actualizar_texto_boton_fondo()
        refrescar = getattr(self, "_refrescar_tabs_fondos", None)
        if callable(refrescar):
            try:
                refrescar()
            except tk.TclError:
                # El diálogo se cerró: su refresco ya no aplica
                self._refrescar_tabs_fondos = None
        self._aplicar_filtros()

    def _actualizar_texto_boton_fondo(self):
        visibles = self._fondos_de_entidades_seleccionadas()
        total = len(visibles)
        seleccionados = [f for f in visibles if self.fondo_vars[f].get()]
        if total == 0:
            texto = "Fondos ▾"
        elif not seleccionados:
            texto = "Fondos: Ninguno ▾"
        elif len(seleccionados) == total:
            texto = "Fondos: Todos ▾"
        else:
            texto = f"Fondos: {len(seleccionados)} de {total} ▾"
        acotados = sum(1 for f in seleccionados if self._series_acotadas(f))
        if acotados:
            texto = texto.replace(" ▾", f" · {acotados} c/serie ▾")
        self.btn_fondo.config(text=texto)

    def _cargar_prefs_json(self):
        try:
            with open(CONFIG_PATH, "r", encoding="utf-8") as f:
                datos = json.load(f)
            return datos if isinstance(datos, dict) else {}
        except (FileNotFoundError, json.JSONDecodeError, OSError, ValueError):
            return {}

    def _cargar_preferencias(self):
        return self._cargar_prefs_json().get("entidades_seleccionadas")

    def _cargar_preferencias_fondos(self):
        return self._cargar_prefs_json().get("fondos_seleccionados")

    def _cargar_preferencias_series_fondo(self):
        guardado = self._cargar_prefs_json().get("serie_por_fondo")
        return guardado if isinstance(guardado, dict) else {}

    def _cargar_catalogo_fondos(self):
        guardado = self._cargar_prefs_json().get("catalogo_fondos")
        return guardado if isinstance(guardado, dict) else {}

    def _guardar_catalogo_fondos(self, catalogo):
        if not catalogo:
            return
        datos = self._cargar_prefs_json()
        combinado = self._cargar_catalogo_fondos()
        combinado.update(catalogo)
        datos["catalogo_fondos"] = combinado
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(datos, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

    def _guardar_preferencias(self):
        datos = self._cargar_prefs_json()
        # Sin datos cargados no hay nada que guardar: no se pisa lo que ya está
        # (por ejemplo una plantilla cargada antes de descargar)
        if self.entidad_vars:
            datos["entidades_seleccionadas"] = [n for n, v in self.entidad_vars.items() if v.get()]
        if self.fondo_vars:
            datos["fondos_seleccionados"] = [f for f, v in self.fondo_vars.items() if v.get()]
            datos["serie_por_fondo"] = {f: sorted(self._series_marcadas(f))
                                        for f in self.fondo_series if self._series_acotadas(f)}
            datos["entidad_por_fondo"] = {f: self.fondo_entidad.get(f, "")
                                          for f, v in self.fondo_vars.items() if v.get()}
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(datos, f, ensure_ascii=False, indent=2)
        except OSError:
            pass

    def _df_entidad_fondo(self):
        """Base filtrada por entidad y por cada fondo con la serie que tenga elegida."""
        if self.df is None:
            return pd.DataFrame()
        resultado = self.df

        if "NOM_ADM" in resultado.columns and self.entidad_vars:
            seleccionadas = [nombre for nombre, v in self.entidad_vars.items() if v.get()]
            if len(seleccionadas) < len(self.entidad_vars):
                resultado = resultado[resultado["NOM_ADM"].isin(seleccionadas)]

        if "RUN_FM" in resultado.columns and self.fondo_vars:
            elegidos = [f for f, v in self.fondo_vars.items() if v.get()]
            acotados = {f: self._series_marcadas(f) for f in elegidos if self._series_acotadas(f)}

            if len(elegidos) < len(self.fondo_vars) or acotados:
                runs = resultado["RUN_FM"].astype(str)
                mascara = runs.isin(elegidos)
                if acotados and "SERIE" in resultado.columns:
                    series = resultado["SERIE"].astype(str)
                    # De un fondo acotado solo sobreviven sus series marcadas
                    for f, marcadas in acotados.items():
                        mascara &= ~((runs == f) & (~series.isin(marcadas)))
                resultado = resultado[mascara]

        return resultado

    def _filtrar_df(self):
        """Aplica entidad + fondo/serie + rango de fechas + filtros de columna + orden."""
        if self.df is None:
            return pd.DataFrame()
        resultado = self._df_entidad_fondo()

        try:
            # Se leen las StringVar, no el widget: dentro del callback del trace
            # el Entry todavía no reflejó el valor nuevo y el filtro se saltaba
            fecha_desde = self.var_fecha_desde.get()
            fecha_hasta = self.var_fecha_hasta.get()
            if "FECHA_INF" in resultado.columns:
                parsed = pd.to_datetime(resultado["FECHA_INF"].astype(str), format="%Y%m%d", errors="coerce")
                resultado = resultado[(parsed >= pd.to_datetime(fecha_desde)) &
                                      (parsed <= pd.to_datetime(fecha_hasta))]
        except Exception:
            pass

        for col_id in self.filtros_columna:
            texto = self.filtro_lectores[col_id]().strip().lower()
            if not texto:
                continue
            campo = COL_CAMPO[col_id]
            if campo not in resultado.columns:
                continue
            if col_id == "fecha":
                texto = texto.replace("-", "")
            serie = resultado[campo].astype(str).str.lower()
            resultado = resultado[serie.str.contains(texto, regex=False, na=False)]

        if self._sort_col:
            campo = COL_CAMPO[self._sort_col]
            if campo in resultado.columns:
                resultado = resultado.sort_values(campo, ascending=self._sort_asc, kind="stable")

        return resultado

    def _aplicar_filtros(self, *_):
        if self.df is None:
            return
        self._llenar_tabla(self._filtrar_df())
        if self.fondo_vars and not any(v.get() for v in self.fondo_vars.values()):
            self.label_seleccion.config(
                text="Ningún fondo marcado — elígelos en el botón «Fondos»")

    def _ajustar_ancho_entidad(self, df):
        """Entidad ocupa justo lo que mide su nombre más largo visible; el resto va a Nombre Fondo."""
        if "NOM_ADM" not in df.columns or df.empty:
            return
        import tkinter.font as tkfont
        fuente = tkfont.Font(font=FONT_NORMAL)
        textos = [str(t) for t in df["NOM_ADM"].dropna().unique()]
        if not textos:
            return
        ancho = max(fuente.measure(t) for t in textos)
        ancho = max(ancho, fuente.measure(COL_ETIQUETA["entidad"])) + 26
        self.tree.column("entidad", width=ancho, minwidth=140)
        self._reposicionar_filtros_columna()

    def _llenar_tabla(self, df):
        self._sel_ancla = None
        self._sel_filas = []
        self._sel_cols = []
        self._actualizar_label_seleccion()
        self._ajustar_ancho_entidad(df)

        for item in self.tree.get_children():
            self.tree.delete(item)

        # Llenar tabla (máximo 1000 filas)
        for _, row in df.head(1000).iterrows():
            fecha = str(row.get('FECHA_INF', ''))
            fecha_fmt = f"{fecha[:4]}-{fecha[4:6]}-{fecha[6:]}" if len(fecha) == 8 else fecha
            valores = (
                fecha_fmt,
                row.get('NOM_ADM', ''),
                row.get('RUN_FM', ''),
                row.get('NOMBRE_FM', ''),
                row.get('SERIE', ''),
                f"{_num(row.get('VALOR_CUOTA')):,.2f}",
            )
            self.tree.insert("", tk.END, values=valores)

    def exportar_excel(self):
        if self.df is None or len(self.df) == 0:
            messagebox.showwarning("Advertencia", "Descarga datos primero")
            return

        df_filtrado = self._filtrar_df()
        if len(df_filtrado) == 0:
            messagebox.showwarning("Advertencia", "No hay datos para exportar con los filtros aplicados")
            return

        if "NOM_ADM" not in df_filtrado.columns:
            messagebox.showerror("Error", "Los datos no traen la columna de entidad (NOM_ADM)")
            return

        hoy = datetime.now().strftime("%Y-%m-%d")
        seleccionadas = sum(1 for v in self.entidad_vars.values() if v.get())
        nombre_default = f"{hoy}_{seleccionadas}-entidades"

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=nombre_default + ".xlsx",
            filetypes=[("Excel", "*.xlsx")]
        )

        if archivo:
            if not archivo.lower().endswith(".xlsx"):
                archivo += ".xlsx"
            try:
                entidades = self._escribir_excel_por_entidad(df_filtrado, archivo)
                messagebox.showinfo(
                    "Éxito",
                    f"Archivo guardado: {archivo}\n\n"
                    f"{entidades} entidad(es), {len(df_filtrado)} fila(s) en total.\n\n"
                    "Se abrirá la carpeta que lo contiene.")
                _abrir_carpeta_de(archivo)
            except Exception as e:
                messagebox.showerror("Error", str(e))

    def _filas_plantilla_guardada(self):
        """Reconstruye la plantilla desde las preferencias, sin necesitar datos cargados."""
        prefs = self._cargar_prefs_json()
        fondos = prefs.get("fondos_seleccionados") or []
        series = prefs.get("serie_por_fondo") or {}
        entidades = prefs.get("entidad_por_fondo") or {}
        nombres = prefs.get("catalogo_fondos") or {}
        filas = []
        for run in fondos:
            elegidas = series.get(run)
            filas.append({
                "Entidad": entidades.get(run, ""),
                "N° Fondo": run,
                "Nombre Fondo": nombres.get(run, ""),
                "Serie": ", ".join(elegidas) if isinstance(elegidas, list) and elegidas else TODAS,
            })
        return filas

    def descargar_plantilla(self):
        """Genera un Excel con Entidad / N° Fondo / Nombre Fondo / Serie para editar y recargar."""
        if self.fondos_disponibles:
            filas = [{
                "Entidad": self.fondo_entidad.get(f, ""),
                "N° Fondo": f,
                "Nombre Fondo": self.fondo_nombre.get(f, ""),
                "Serie": (TODAS if not self._series_acotadas(f)
                          else ", ".join(sorted(self._series_marcadas(f)))),
            } for f in self.fondos_disponibles if self.fondo_vars[f].get()]
        else:
            # Sin datos cargados se usa lo guardado en preferencias
            filas = self._filas_plantilla_guardada()

        if not filas:
            if not messagebox.askyesno(
                    "Plantilla",
                    "No hay fondos seleccionados ni preferencias guardadas.\n\n"
                    "¿Generar una plantilla vacía con solo los encabezados para "
                    "llenarla a mano?"):
                return
            filas = []

        plantilla = pd.DataFrame(filas, columns=COLS_PLANTILLA)
        if not plantilla.empty:
            plantilla = plantilla.sort_values(["Entidad", "N° Fondo"], key=lambda c:
                                              c.map(_clave_fondo) if c.name == "N° Fondo" else c)

        archivo = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            initialfile=f"plantilla_fondos_{datetime.now():%Y-%m-%d}.xlsx",
            filetypes=[("Excel", "*.xlsx")])
        if not archivo:
            return
        if not archivo.lower().endswith(".xlsx"):
            archivo += ".xlsx"

        try:
            self._escribir_excel(plantilla, archivo, hoja="Plantilla")
            self.label_plantilla.config(text=f"Plantilla generada: {len(plantilla)} fondos",
                                        foreground="#4CAF50")
            messagebox.showinfo("Plantilla",
                                f"Plantilla guardada con {len(plantilla)} fondos:\n{archivo}\n\n"
                                "Edítala y vuelve a cargarla para fijar qué se muestra por defecto.\n"
                                "En 'Serie' escribe 'Todas' o los nombres separados por coma.")
            _abrir_carpeta_de(archivo)
        except Exception as e:
            messagebox.showerror("Error", str(e))

    def _leer_plantilla(self, ruta):
        """Lee el Excel y devuelve {run: {entidad, series}}; None si no es válida."""
        try:
            plantilla = pd.read_excel(ruta)
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo leer la plantilla:\n{e}")
            return None

        if "N° Fondo" not in plantilla.columns:
            messagebox.showerror(
                "Plantilla inválida",
                "La plantilla debe tener al menos la columna 'N° Fondo'.\n"
                f"Columnas esperadas: {', '.join(COLS_PLANTILLA)}")
            return None

        pedidos = {}
        for _, fila in plantilla.iterrows():
            run = str(fila["N° Fondo"]).strip()
            if run.endswith(".0"):        # Excel puede leer el número como float
                run = run[:-2]
            if not run or run.lower() == "nan":
                continue
            entidad = str(fila.get("Entidad", "") or "").strip()
            if entidad.lower() == "nan":
                entidad = ""
            celda = str(fila.get("Serie", TODAS) or TODAS).strip()
            if celda.lower() in ("nan", "", TODAS.lower()):
                series = []                # vacío = todas las series del fondo
            else:
                # Se aceptan coma, punto y coma o salto de línea como separadores
                series = [s.strip() for s in re.split(r"[,;\n]", celda) if s.strip()]

            # Un mismo fondo puede venir repetido, una fila por serie: se acumulan
            # en vez de sobrescribirse (antes solo sobrevivía la última fila).
            entrada = pedidos.setdefault(run, {"entidad": entidad, "series": [],
                                               "todas": False})
            if entidad and not entrada["entidad"]:
                entrada["entidad"] = entidad
            if series:
                for serie in series:
                    if serie not in entrada["series"]:
                        entrada["series"].append(serie)
            else:
                entrada["todas"] = True    # alguna fila pide el fondo completo

        # "Todas" manda: si el fondo aparece sin acotar en alguna fila, va completo
        for entrada in pedidos.values():
            if entrada.pop("todas", False):
                entrada["series"] = []
        return pedidos

    def cargar_plantilla(self):
        """Aplica una plantilla. Si aún no hay datos, la deja guardada para aplicarla al descargar."""
        archivo = filedialog.askopenfilename(filetypes=[("Excel", "*.xlsx"), ("Todos", "*.*")])
        if not archivo:
            return

        pedidos = self._leer_plantilla(archivo)
        if pedidos is None:
            return
        if not pedidos:
            messagebox.showwarning("Plantilla", "La plantilla no tiene ninguna fila con N° Fondo")
            return

        if not self.fondos_disponibles:
            # Sin datos todavía: se guarda y se aplicará sola tras la descarga
            self._guardar_plantilla_pendiente(pedidos)
            self.label_plantilla.config(
                text=f"Plantilla guardada: {len(pedidos)} fondos (se aplicará al descargar)",
                foreground="#5B9BD5")
            messagebox.showinfo(
                "Plantilla",
                f"Se guardaron {len(pedidos)} fondos como preferencia.\n\n"
                "Todavía no hay datos cargados, así que la selección se aplicará "
                "automáticamente en cuanto descargues la cartola.")
            return

        aplicados, series_aplicadas, desconocidos = self._aplicar_plantilla(pedidos)
        if not aplicados:
            messagebox.showwarning("Plantilla",
                                   "Ningún fondo de la plantilla coincide con los datos cargados")
            return

        self.label_plantilla.config(text=f"Plantilla aplicada: {len(aplicados)} fondos",
                                    foreground="#4CAF50")
        aviso = f"Se aplicaron {len(aplicados)} fondos y {series_aplicadas} series."
        if desconocidos:
            muestra = ", ".join(desconocidos[:8])
            extra = f" y {len(desconocidos) - 8} más" if len(desconocidos) > 8 else ""
            aviso += f"\n\nNo se reconocieron: {muestra}{extra}"
        messagebox.showinfo("Plantilla", aviso)

    def _guardar_plantilla_pendiente(self, pedidos):
        """Escribe la plantilla en las preferencias para aplicarla cuando lleguen datos."""
        datos = self._cargar_prefs_json()
        datos["fondos_seleccionados"] = list(pedidos)
        datos["serie_por_fondo"] = {run: info["series"]
                                    for run, info in pedidos.items() if info["series"]}
        datos["entidad_por_fondo"] = {run: info["entidad"] for run, info in pedidos.items()}
        entidades = sorted({info["entidad"] for info in pedidos.values() if info["entidad"]})
        if entidades:
            datos["entidades_seleccionadas"] = entidades
        try:
            with open(CONFIG_PATH, "w", encoding="utf-8") as f:
                json.dump(datos, f, ensure_ascii=False, indent=2)
        except OSError as e:
            messagebox.showerror("Error", f"No se pudo guardar la plantilla: {e}")

    def _aplicar_plantilla(self, pedidos):
        """Casa la plantilla con los datos cargados por Entidad + N° Fondo + Serie."""
        aplicados, series_aplicadas, desconocidos = [], 0, []
        series_por_run = {}

        for run, info in pedidos.items():
            if run not in self.fondo_vars:
                desconocidos.append(run)
                continue
            # La entidad, si viene en la plantilla, debe coincidir con la real
            entidad_real = self.fondo_entidad.get(run, "")
            if info["entidad"] and info["entidad"].strip().lower() != entidad_real.strip().lower():
                desconocidos.append(f"{run} (entidad no coincide)")
                continue
            aplicados.append(run)

            # El nombre de la serie se compara sin distinguir mayúsculas ni espacios,
            # pero se guarda tal como viene en la cartola
            disponibles = self.series_por_fondo.get(run, [])
            equivalencias = {str(s).strip().upper(): s for s in disponibles}
            validas, invalidas = set(), []
            for pedida in info["series"]:
                real = equivalencias.get(str(pedida).strip().upper())
                if real is not None:
                    validas.add(real)
                else:
                    invalidas.append(pedida)
            if validas:
                series_por_run[run] = validas
                series_aplicadas += len(validas)
            if invalidas:
                desconocidos.append(f"{run} (serie(s) {', '.join(invalidas)})")

        if not aplicados:
            return [], 0, desconocidos

        for f, var in self.fondo_vars.items():
            var.set(f in aplicados)
        for f in self.fondo_series:
            self.fondo_series[f] = series_por_run.get(f, set(self.series_por_fondo.get(f, [])))

        entidades = {self.fondo_entidad.get(f, "") for f in aplicados}
        for nombre, var in self.entidad_vars.items():
            var.set(nombre in entidades)

        self._guardar_preferencias()
        self._actualizar_texto_boton_entidad()
        self._actualizar_texto_boton_fondo()
        self._aplicar_filtros()
        return aplicados, series_aplicadas, desconocidos

    def _escribir_excel(self, df_export, archivo, hoja="Fondos"):
        """Escribe una tabla plana ajustando el ancho de cada columna al contenido."""
        from openpyxl.utils import get_column_letter

        with pd.ExcelWriter(archivo, engine="openpyxl") as writer:
            df_export.to_excel(writer, index=False, sheet_name=hoja)
            ws = writer.sheets[hoja]
            for idx, titulo in enumerate(df_export.columns, start=1):
                letra = get_column_letter(idx)
                if titulo == "Valor Cuota":
                    muestra = [f"{_num(v):,.2f}" for v in df_export[titulo].head(500)]
                    for celda in ws[letra][1:]:
                        celda.number_format = "#,##0.00"
                else:
                    muestra = [str(v) for v in df_export[titulo].head(500)]
                ancho = max([len(str(titulo))] + [len(v) for v in muestra])
                ws.column_dimensions[letra].width = min(ancho + 4, 60)

    def _escribir_excel_por_entidad(self, df, archivo):
        """Escribe el Excel agrupado: cada entidad encabeza sus fondos, separadas por 2 filas."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        campos = [c for c, _ in CAMPOS_EXPORT if c in df.columns]
        titulos = [t for c, t in CAMPOS_EXPORT if c in df.columns]

        wb = Workbook()
        ws = wb.active
        ws.title = "Fondos"

        fuente_entidad = Font(bold=True, size=12, color="FFFFFF")
        relleno_entidad = PatternFill("solid", fgColor="5B9BD5")
        fuente_cab = Font(bold=True)
        relleno_cab = PatternFill("solid", fgColor="D9E2F3")

        anchos = [len(t) for t in titulos]
        fila = 1
        entidades = 0

        for entidad, grupo in df.groupby("NOM_ADM", sort=True):
            entidades += 1
            celda = ws.cell(row=fila, column=1, value=str(entidad))
            celda.font = fuente_entidad
            celda.fill = relleno_entidad
            ws.merge_cells(start_row=fila, start_column=1,
                           end_row=fila, end_column=max(len(titulos), 1))
            fila += 1

            for idx, titulo in enumerate(titulos, start=1):
                celda = ws.cell(row=fila, column=idx, value=titulo)
                celda.font = fuente_cab
                celda.fill = relleno_cab
            fila += 1

            if "RUN_FM" in grupo.columns:
                grupo = grupo.sort_values("RUN_FM", key=lambda s: s.map(_clave_fondo))

            for _, registro in grupo.iterrows():
                for idx, campo in enumerate(campos, start=1):
                    valor = registro.get(campo, "")
                    if campo == "VALOR_CUOTA":
                        valor = _num(valor)
                        celda = ws.cell(row=fila, column=idx, value=valor)
                        celda.number_format = "#,##0.00"
                        largo = len(f"{valor:,.2f}")
                    else:
                        if pd.isna(valor):
                            valor = ""
                        elif campo == "FECHA_INF":
                            texto = str(valor)
                            if len(texto) == 8:
                                valor = f"{texto[:4]}-{texto[4:6]}-{texto[6:]}"
                        celda = ws.cell(row=fila, column=idx, value=valor)
                        largo = len(str(valor))
                    anchos[idx - 1] = max(anchos[idx - 1], largo)
                fila += 1

            fila += 2   # dos filas en blanco antes de la siguiente entidad

        for idx, ancho in enumerate(anchos, start=1):
            ws.column_dimensions[get_column_letter(idx)].width = min(ancho + 4, 60)

        wb.save(archivo)
        return entidades


if __name__ == "__main__":
    root = tk.Tk()
    app = FondosMutuosApp(root)
    root.mainloop()
